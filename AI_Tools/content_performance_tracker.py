"""
Content Performance Tracker Excel
===================================
Tracks reel/post performance per client — views, reach, saves, DMs, leads generated.

Output: 04_AgentKPI/content_performance.xlsx
Usage:  py -3 AI_Tools/create_content_performance.py
"""

import sys
from pathlib import Path
from datetime import date, timedelta

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import LineChart, BarChart, Reference
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("pip install openpyxl"); sys.exit(1)

OUTPUT = Path(__file__).parent.parent / "04_AgentKPI" / "content_performance.xlsx"
TODAY  = date.today()

C = {
    "nav":     "1A1A2E", "card": "16213E", "white": "FFFFFF",
    "lgray":   "F4F4F4", "mgray": "AAAAAA", "dgray": "444444",
    "consult": "4472C4", "re":    "70AD47", "gold":  "FFD700",
    "green":   "C6EFCE", "red":   "FF6B6B", "orange":"FFD580",
}
CLIENTS = ["Fadli", "Ruth", "Mathew Jordan"]

def _fill(h): return PatternFill(start_color=h, end_color=h, fill_type="solid")
def _bdr(c="CCCCCC"):
    s = Side(style="thin", color=c)
    return Border(left=s, right=s, top=s, bottom=s)
def _font(c="000000", sz=11, bold=False, italic=False):
    return Font(name="Calibri", size=sz, bold=bold, italic=italic, color=c)
def _align(h="center", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def s(cell, val=None, bg=None, fc="000000", sz=11, bold=False, h="center", bdr=True, italic=False):
    if val is not None: cell.value = val
    if bg: cell.fill = _fill(bg)
    cell.font = _font(fc, sz, bold, italic)
    cell.alignment = _align(h)
    if bdr: cell.border = _bdr()


def build_log(ws):
    """Main performance log — one row per post"""
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = C["consult"]

    COLS = [
        ("A",3), ("B",14), ("C",16), ("D",16), ("E",14), ("F",18),
        ("G",12), ("H",12), ("I",12), ("J",12), ("K",14), ("L",14),
        ("M",14), ("N",14), ("O",16), ("P",3),
    ]
    for col, w in COLS:
        ws.column_dimensions[col].width = w

    ws.row_dimensions[2].height = 50
    ws.merge_cells("B2:O2")
    s(ws["B2"], "📱  CONTENT PERFORMANCE LOG — All Clients",
      bg=C["nav"], fc=C["gold"], sz=18, bold=True)

    ws.row_dimensions[3].height = 22
    ws.merge_cells("B3:O3")
    s(ws["B3"],
      '=COUNTA(B5:B504)&" posts tracked  |  Avg views: "&TEXT(IFERROR(AVERAGE(G5:G504),0),"#,##0")&"  |  Total DMs: "&TEXT(SUM(K5:K504),"#,##0")',
      bg=C["card"], fc=C["gold"], sz=10)

    hdrs = ["Post Date", "Client", "Platform", "Format", "Hook (first 5 words)",
            "Views", "Reach", "Saves", "Shares", "Comments", "DMs from Post",
            "Leads Generated", "CTR Score (1-10)", "Notes"]
    ws.row_dimensions[4].height = 40
    for i, h in enumerate(hdrs):
        c = ws.cell(row=4, column=2+i)
        s(c, h, bg=C["consult"], fc=C["white"], sz=9, bold=True)

    for r in range(5, 505):
        ws.row_dimensions[r].height = 20
        row_bg = "FFFFFF" if r % 2 == 0 else C["lgray"]
        for col in range(2, 16):
            c = ws.cell(row=r, column=col)
            c.border = _bdr("DDDDDD")
            c.fill = _fill(row_bg)
            c.font = _font("000000", 10)
            c.alignment = _align("left" if col in [2,3,4,5,6,15] else "center")

        ws.cell(row=r, column=2).number_format  = "DD-MMM-YY"
        for num_col in [7,8,9,10,11,12,13]:
            ws.cell(row=r, column=num_col).number_format = "#,##0"

    # Dropdowns
    dv_client = DataValidation(type="list",
        formula1='"' + ",".join(CLIENTS) + '"', allow_blank=True)
    dv_platform = DataValidation(type="list",
        formula1='"IG Reel,IG Story,TikTok,YouTube Short,Threads,YouTube Long"',
        allow_blank=True)
    dv_format = DataValidation(type="list",
        formula1='"Talking Head,B-Roll + VO,Text Overlay,Carousel,Story Sequence,Mixed"',
        allow_blank=True)
    ws.add_data_validation(dv_client);   dv_client.sqref   = "C5:C504"
    ws.add_data_validation(dv_platform); dv_platform.sqref = "D5:D504"
    ws.add_data_validation(dv_format);   dv_format.sqref   = "E5:E504"

    # CTR Score color scale
    ws.conditional_formatting.add("N5:N504",
        ColorScaleRule(
            start_type="num", start_value=1,  start_color=C["red"],
            mid_type="num",   mid_value=5,    mid_color=C["orange"],
            end_type="num",   end_value=10,   end_color=C["green"],
        ))

    ws.freeze_panes = "B5"


def build_summary(ws):
    """Per-client summary with charts"""
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = C["card"]

    for col, w in [("A",3),("B",22),("C",16),("D",16),("E",16),
                   ("F",16),("G",16),("H",16),("I",3)]:
        ws.column_dimensions[col].width = w

    ws.row_dimensions[2].height = 50
    ws.merge_cells("B2:H2")
    s(ws["B2"], "📊  PERFORMANCE SUMMARY — By Client",
      bg=C["nav"], fc=C["gold"], sz=18, bold=True)

    # Summary table
    ws.row_dimensions[4].height = 26
    for i, h in enumerate(["Client", "Posts", "Avg Views", "Total Saves",
                            "Total DMs", "Leads", "Avg CTR Score"]):
        s(ws.cell(row=4, column=2+i), h, bg=C["nav"], fc=C["white"], sz=10, bold=True)

    for idx, client in enumerate(CLIENTS):
        r = 5 + idx
        ws.row_dimensions[r].height = 26
        row_bg = [C["consult"], C["re"], "7030A0"][idx]

        s(ws.cell(row=r, column=2), client, bg=row_bg, fc=C["white"], sz=11, bold=True)

        # COUNTIF / AVERAGEIF / SUMIF referencing Log sheet
        formulas = [
            f'=COUNTIF(Log!C5:C504,"{client}")',
            f'=IFERROR(AVERAGEIF(Log!C5:C504,"{client}",Log!G5:G504),0)',
            f'=SUMIF(Log!C5:C504,"{client}",Log!I5:I504)',
            f'=SUMIF(Log!C5:C504,"{client}",Log!K5:K504)',
            f'=SUMIF(Log!C5:C504,"{client}",Log!L5:L504)',
            f'=IFERROR(AVERAGEIF(Log!C5:C504,"{client}",Log!N5:N504),0)',
        ]
        for ci, formula in enumerate(formulas):
            c = ws.cell(row=r, column=3+ci)
            c.value = formula
            c.font = _font("000000", 11)
            c.alignment = _align("center")
            c.border = _bdr()
            c.fill = _fill(C["lgray"])
            if ci in [1,2,3,4]:
                c.number_format = "#,##0"
            elif ci == 5:
                c.number_format = "0.0"

    # Chart: Avg Views by client
    ws.row_dimensions[9].height = 8
    bar = BarChart()
    bar.type = "col"
    bar.title = "Average Views by Client"
    bar.y_axis.title = "Views"
    bar.y_axis.numFmt = "#,##0"
    bar.style = 10
    bar.width = 16
    bar.height = 10
    bar.legend = None

    data = Reference(ws, min_col=4, max_col=4, min_row=4, max_row=7)
    cats = Reference(ws, min_col=2, min_row=5, max_row=7)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats)
    ws.add_chart(bar, "B10")

    ws.freeze_panes = "B5"


def main():
    wb = Workbook()
    ws_log = wb.active;           ws_log.title     = "Log"
    ws_sum = wb.create_sheet("Summary")

    print("  Building: Log")
    build_log(ws_log)
    print("  Building: Summary")
    build_summary(ws_sum)

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    print(f"\nSaved: {OUTPUT}")


if __name__ == "__main__":
    main()
