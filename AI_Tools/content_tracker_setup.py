"""
create_content_tracker.py
--------------------------
One-time setup script. Creates VLT_Content/Content_Tracker.xlsx and seeds it
with all existing approved content found in 02_HMN_HUMANFLOW.

Run once:
  py -3 AI_Tools/create_content_tracker.py

After that, use Content_Tracker.xlsx directly:
  - Change Status from "review" → "approved"
  - Run: py -3 AI_Tools/content_approve.py
"""

import re
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

CONTENT_ROOT = Path(__file__).parent.parent / "Content"
TRACKER_PATH = CONTENT_ROOT / "Content_Tracker.xlsx"
HUMANFLOW    = CONTENT_ROOT / "02_HMN_HUMANFLOW"

HEADERS = [
    "ID", "Client", "Month", "Post_Date", "Slug",
    "Type", "Status", "Script_Path", "Final_Path", "Approved_Date", "Notes"
]

# Status fill colors
STATUS_COLORS = {
    "draft":    "FFF2CC",  # yellow
    "review":   "FCE5CD",  # orange
    "approved": "D9EAD3",  # green
    "rendered": "CFE2F3",  # blue
    "posted":   "E8E8E8",  # grey
}

SLUG_RE = re.compile(
    r"^(\d{4}-\d{2}-\d{2})_([a-z\-]+)_(.+)$"
)

def parse_filename(stem: str):
    """Extract (post_date, type, slug) from filename stem."""
    m = SLUG_RE.match(stem)
    if m:
        return m.group(1), m.group(2), stem
    return "", "unknown", stem

def collect_existing() -> list[dict]:
    """Scan 02_HMN_HUMANFLOW for all approved .md files."""
    rows = []
    for md in sorted(HUMANFLOW.rglob("approved/*.md")):
        # Derive client from path: jocons/fadli or real_estate
        parts = md.relative_to(HUMANFLOW).parts
        # parts: e.g. ('jocons', 'fadli', 'projects', '2026-03', 'approved', 'file.md')
        if len(parts) >= 2 and parts[0] == "jocons":
            client = parts[1] if len(parts) > 1 else parts[0]
        else:
            client = parts[0]

        # Month from path
        month = ""
        for p in parts:
            if re.match(r"\d{4}-\d{2}$", p):
                month = p
                break

        post_date, content_type, slug = parse_filename(md.stem)
        final_rel = str(md.relative_to(CONTENT_ROOT))

        rows.append({
            "client":        client,
            "month":         month or post_date[:7] if post_date else "",
            "post_date":     post_date,
            "slug":          slug,
            "type":          content_type,
            "status":        "approved",
            "script_path":   "",
            "final_path":    final_rel,
            "approved_date": "",
            "notes":         "seeded from existing approved folder",
        })
    return rows

def style_header(ws):
    header_fill = PatternFill("solid", fgColor="434343")
    header_font = Font(bold=True, color="FFFFFF")
    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

def set_col_widths(ws):
    widths = [5, 14, 10, 12, 45, 12, 10, 55, 55, 14, 25]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def write_row(ws, row_num: int, row_id: int, data: dict):
    values = [
        row_id,
        data["client"],
        data["month"],
        data["post_date"],
        data["slug"],
        data["type"],
        data["status"],
        data["script_path"],
        data["final_path"],
        data["approved_date"],
        data["notes"],
    ]
    status = data["status"].lower()
    fill_color = STATUS_COLORS.get(status, "FFFFFF")
    fill = PatternFill("solid", fgColor=fill_color)

    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.fill = fill
        cell.alignment = Alignment(wrap_text=False)

def main():
    if TRACKER_PATH.exists():
        print(f"Tracker already exists at {TRACKER_PATH}")
        print("Delete it manually if you want to regenerate from scratch.")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Content"
    ws.freeze_panes = "A2"

    style_header(ws)
    set_col_widths(ws)

    existing = collect_existing()
    print(f"Found {len(existing)} existing approved files to seed.")

    for i, row_data in enumerate(existing, 1):
        write_row(ws, i + 1, i, row_data)

    # Add a blank template row at the end for new entries
    next_id = len(existing) + 1
    template = {
        "client": "fadli",
        "month": "2026-03",
        "post_date": "2026-03-01",
        "slug": "YYYY-MM-DD_type_slug",
        "type": "reel",
        "status": "review",
        "script_path": "03_HMN_REVIEW/fadli/YYYY-MM-DD_type_slug.md",
        "final_path": "",
        "approved_date": "",
        "notes": "← example row, delete or overwrite",
    }
    write_row(ws, next_id + 1, next_id, template)

    wb.save(TRACKER_PATH)
    print(f"Tracker created: {TRACKER_PATH}")
    print(f"Rows seeded: {len(existing)}")
    print("\nNext steps:")
    print("  1. Open Content_Tracker.xlsx in Excel or Obsidian")
    print("  2. When you review a file in 03_HMN_REVIEW/, change its Status to 'approved'")
    print("  3. Run: py -3 AI_Tools/content_approve.py")

if __name__ == "__main__":
    main()
