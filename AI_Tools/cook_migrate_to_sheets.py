#!/usr/bin/env python3
"""
Migrate Data/JO_Consult_Dashboard.xlsx → Google Sheets (exact replica).

Replicates all 5 tabs: CRM, EOD, SetterDash, CashDash, PaymentPlans.
Preserves: data values, formulas, headers, frozen rows, tab colors.

Run:
  python tools/migrate_to_sheets.py

First run opens browser for Google sign-in.
After that, token is cached — no manual refresh ever needed.
Saves sheet ID to .env as GOOGLE_SHEETS_DASHBOARD_ID.
"""

import json
import os
import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl
from dotenv import load_dotenv, set_key

ENV_PATH = Path(__file__).parent.parent / ".env"
XLSX_PATH = Path(__file__).parent.parent / "Data" / "JO_Consult_Dashboard.xlsx"

load_dotenv(ENV_PATH)

# ─── Tab Colors ───────────────────────────────────────────────────────────────

TAB_COLORS = {
    "CRM":          {"red": 0.69, "green": 0.95, "blue": 0.73},   # green
    "EOD":          {"red": 1.0,  "green": 0.85, "blue": 0.66},   # orange
    "SetterDash":   {"red": 0.64, "green": 0.84, "blue": 1.0},    # blue
    "CashDash":     {"red": 1.0,  "green": 0.79, "blue": 0.79},   # red
    "PaymentPlans": {"red": 0.82, "green": 0.75, "blue": 1.0},    # purple
}

# Header background colors (same palette, for row 1 formatting)
HEADER_COLORS = {
    "CRM":          {"red": 0.69, "green": 0.95, "blue": 0.73},
    "EOD":          {"red": 1.0,  "green": 0.85, "blue": 0.66},
    "SetterDash":   {"red": 0.64, "green": 0.84, "blue": 1.0},
    "CashDash":     {"red": 1.0,  "green": 0.79, "blue": 0.79},
    "PaymentPlans": {"red": 0.82, "green": 0.75, "blue": 1.0},
}


# ─── Helpers ──────────────────────────────────────────────────────────────────

def col_to_letter(n: int) -> str:
    """Convert 1-based column index to spreadsheet letter (1→A, 27→AA, 40→AN)."""
    result = ""
    while n > 0:
        n -= 1
        result = chr(n % 26 + ord("A")) + result
        n //= 26
    return result


def to_sheets_value(v):
    """Convert openpyxl cell value to a gspread-safe value."""
    if v is None:
        return ""
    if isinstance(v, str):
        stripped = v.strip()
        if stripped.startswith("="):
            return stripped  # formula — pass through as-is
        return stripped
    if isinstance(v, datetime):
        return v.strftime("%d/%m/%Y")
    if isinstance(v, date):
        return v.strftime("%d/%m/%Y")
    if isinstance(v, bool):
        return v  # gspread handles native booleans
    return v


def last_data_row(ws) -> int:
    """Find the last row containing at least one non-None cell."""
    for r in range(ws.max_row, 0, -1):
        if any(ws.cell(r, c).value is not None for c in range(1, ws.max_column + 1)):
            return r
    return 1


# ─── OAuth Setup ──────────────────────────────────────────────────────────────

def setup_oauth():
    """Write gspread credentials.json from .env vars."""
    client_id = os.getenv("GOOGLE_CLIENT_ID")
    client_secret = os.getenv("GOOGLE_CLIENT_SECRET")

    if not client_id or not client_secret:
        print("\nERROR: GOOGLE_CLIENT_ID / GOOGLE_CLIENT_SECRET missing from .env")
        sys.exit(1)

    try:
        import gspread.auth
        creds_dir = Path(gspread.auth.DEFAULT_CONFIG_DIR)
    except Exception:
        creds_dir = Path.home() / ".config" / "gspread"
    creds_dir.mkdir(parents=True, exist_ok=True)

    creds_file = creds_dir / "credentials.json"
    with open(creds_file, "w") as f:
        json.dump({
            "installed": {
                "client_id": client_id,
                "client_secret": client_secret,
                "redirect_uris": ["http://localhost"],
                "auth_uri": "https://accounts.google.com/o/oauth2/auth",
                "token_uri": "https://oauth2.googleapis.com/token",
            }
        }, f)

    return creds_file


# ─── Main Migration ───────────────────────────────────────────────────────────

def migrate():
    try:
        import gspread
    except ImportError:
        print("ERROR: gspread not installed. Run: pip install -r tools/requirements.txt")
        sys.exit(1)

    if not XLSX_PATH.exists():
        print(f"ERROR: Source file not found: {XLSX_PATH}")
        sys.exit(1)

    # Check for existing sheet
    existing_id = os.getenv("GOOGLE_SHEETS_DASHBOARD_ID")
    if existing_id:
        confirm = input(f"\nSheet ID already in .env ({existing_id}). Recreate? [y/N]: ")
        if confirm.lower() != "y":
            print("Aborted.")
            return

    # Load workbook
    print(f"\nReading: {XLSX_PATH.name}")
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=False)
    print(f"Tabs: {wb.sheetnames}")

    # Setup OAuth and authenticate
    setup_oauth()
    print("\nOpening browser for Google sign-in (token cached after first use)...")
    gc = gspread.oauth()

    # Create spreadsheet
    title = "JO Consult — Dashboard"
    print(f"\nCreating spreadsheet: '{title}'...")
    sh = gc.create(title)

    worksheets_created = []
    first = True

    for sheet_name in wb.sheetnames:
        ws_xl = wb[sheet_name]
        max_col = ws_xl.max_column
        end_row = last_data_row(ws_xl)

        # Create the worksheet
        if first:
            ws_gs = sh.sheet1
            ws_gs.update_title(sheet_name)
            first = False
        else:
            ws_gs = sh.add_worksheet(
                title=sheet_name,
                rows=max(end_row + 20, 200),
                cols=max(max_col + 2, 40),
            )

        worksheets_created.append(ws_gs)

        # Build data matrix (only up to last data row, preserving row positions)
        data = []
        for r in range(1, end_row + 1):
            row = [to_sheets_value(ws_xl.cell(r, c).value) for c in range(1, max_col + 1)]
            data.append(row)

        # Write all at once with USER_ENTERED so formulas evaluate
        # gspread 6.x: range_name is first arg
        if data:
            ws_gs.update("A1", data, value_input_option="USER_ENTERED")

        # Bold + colored header (row 1)
        col_letter = col_to_letter(max_col)
        color = HEADER_COLORS.get(sheet_name, {"red": 0.9, "green": 0.9, "blue": 0.9})
        ws_gs.format(f"A1:{col_letter}1", {
            "textFormat": {"bold": True, "fontSize": 10},
            "backgroundColor": color,
        })

        # Freeze header row
        ws_gs.freeze(rows=1)

        print(f"  ✓ {sheet_name:<15} ({end_row} rows × {max_col} cols)")

    # Set tab colors via batch API
    tab_color_requests = []
    for i, sheet_name in enumerate(wb.sheetnames):
        color = TAB_COLORS.get(sheet_name, {"red": 0.9, "green": 0.9, "blue": 0.9})
        tab_color_requests.append({
            "updateSheetProperties": {
                "properties": {
                    "sheetId": worksheets_created[i].id,
                    "tabColor": color,
                },
                "fields": "tabColor",
            }
        })

    if tab_color_requests:
        sh.batch_update({"requests": tab_color_requests})

    # Save and report
    url = f"https://docs.google.com/spreadsheets/d/{sh.id}"
    set_key(str(ENV_PATH), "GOOGLE_SHEETS_DASHBOARD_ID", sh.id)

    print(f"\nDone.")
    print(f"  URL: {url}")
    print(f"  ID saved to .env as GOOGLE_SHEETS_DASHBOARD_ID")
    print(f"\nNext: Share the sheet with your n8n service account as Editor:")
    print(f"  n8nreader@festive-kayak-459112-e0.iam.gserviceaccount.com")
    print(f"\nAfter sharing, n8n can read/write using the service account JSON key.")

    return sh.id


# ─── Entry Point ─────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("JO Consult — Dashboard Migration to Google Sheets")
    print("=" * 50)
    migrate()
