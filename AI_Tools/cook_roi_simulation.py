#!/usr/bin/env python3
"""
ROI Simulation Generator — Create professional property investment simulations.

Usage:
    py -3 AI_Tools/create_roi_simulation.py --client "Andi Pratama" --property "SAVASA 2BR" --price 1500000000 --output "path/to/ROI.xlsx"
"""

import argparse
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from pathlib import Path

def create_roi_simulation(file_path, client_name, property_name, price):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ROI Simulation"

    # Header Styling
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    header_font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")

    ws["A1"] = f"ROI Simulation for {client_name}"
    ws["A1"].font = header_font
    ws["A1"].fill = header_fill
    ws.merge_cells("A1:D1")
    ws["A1"].alignment = Alignment(horizontal="center")

    data = [
        ("Property", property_name),
        ("Purchase Price", price),
        ("", ""),
        ("Yearly Appreciation (Est. 5%)", price * 0.05),
        ("Monthly Rental Income (Est.)", 8000000),
        ("Yearly Rental Income", 96000000),
        ("Maintenance & Tax (Yearly)", price * 0.01),
        ("", ""),
        ("Net Yearly Profit (Rental + Appr.)", (price * 0.05) + 96000000 - (price * 0.01)),
        ("Total ROI % (Year 1)", ((price * 0.05) + 96000000 - (price * 0.01)) / price * 100)
    ]

    for row_idx, (label, value) in enumerate(data, 3):
        ws.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row_idx, column=2, value=value)
        if isinstance(value, (int, float)) and label != "Total ROI % (Year 1)":
            ws.cell(row=row_idx, column=2).number_format = '#,##0'

    # Adjust width
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20

    wb.save(file_path)
    print(f"ROI Simulation created: {file_path}")

def main():
    parser = argparse.ArgumentParser(description="Generate ROI Simulation Excel")
    parser.add_argument("--client", required=True)
    parser.add_argument("--property", required=True)
    parser.add_argument("--price", type=float, required=True)
    parser.add_argument("--output", required=True)
    
    args = parser.parse_args()
    create_roi_simulation(args.output, args.client, args.property, args.price)

if __name__ == "__main__":
    main()
