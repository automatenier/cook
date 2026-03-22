"""
RE CRM Setup — creates Real Estate CRM tracker at:
01 HMN__Command/11_Real_Estate/Dash/RE_CRM.xlsx

Sheets:
  1. Pipeline    — one row per prospect, all qualification fields
  2. Revenue_Log — closed deals + commission tracking
  3. Lead_Source — source attribution summary

Run: py -3 AI_Tools/re_crm_setup.py
"""

import os
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("Installing openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl -q")
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

# ── Paths ──────────────────────────────────────────────────────────────────────
BASE = Path(__file__).resolve().parent.parent
OUTPUT = BASE / "01 HMN__Command" / "11_Real_Estate" / "Dash" / "RE_CRM.xlsx"
OUTPUT.parent.mkdir(parents=True, exist_ok=True)

# ── Palette ────────────────────────────────────────────────────────────────────
C_HEADER_BG   = "1A1A2E"   # dark navy
C_HEADER_FG   = "FFFFFF"
C_HOT         = "FF4D4D"
C_WARM        = "FF9900"
C_COLD        = "4D9FFF"
C_WON         = "27AE60"
C_LOST        = "95A5A6"
C_ALT_ROW     = "F5F7FA"
C_BORDER      = "D0D5DD"
C_ACCENT      = "0A66C2"

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, size=size, name="Calibri")

def _border():
    side = Side(style="thin", color=C_BORDER)
    return Border(left=side, right=side, top=side, bottom=side)

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def _style_header_row(ws, row, col_count):
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = _fill(C_HEADER_BG)
        cell.font = _font(bold=True, color=C_HEADER_FG, size=10)
        cell.alignment = _center()
        cell.border = _border()

def _style_data_row(ws, row, col_count, alt=False):
    bg = C_ALT_ROW if alt else "FFFFFF"
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        if not cell.fill or cell.fill.fgColor.rgb in ("00000000", "FFFFFFFF", C_ALT_ROW):
            cell.fill = _fill(bg)
        cell.font = _font(size=10)
        cell.alignment = _left()
        cell.border = _border()

def build_pipeline_sheet(wb):
    ws = wb.active
    ws.title = "Pipeline"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    # Column definitions: (header, width)
    cols = [
        ("Name",          18),
        ("IG Handle",     16),
        ("WhatsApp",      16),
        ("Source",        16),
        ("Status",        18),
        ("Temp",           8),
        ("Intent",        12),
        ("Budget Min (M)", 13),
        ("Budget Max (M)", 13),
        ("Area",          16),
        ("Property Type", 14),
        ("Financing",     12),
        ("Timeline",      14),
        ("Created",       12),
        ("Last Contact",  12),
        ("Next Action",   24),
        ("Notes",         30),
    ]

    # Write headers
    for i, (header, width) in enumerate(cols, 1):
        cell = ws.cell(row=1, column=i, value=header)
        ws.column_dimensions[get_column_letter(i)].width = width

    _style_header_row(ws, 1, len(cols))
    ws.row_dimensions[1].height = 28

    # Seed with 1 example row
    example = [
        "Budi Santoso", "@budisantoso", "+6281234567890",
        "Referral", "response", "WARM",
        "Invest", 1.2, 1.8,
        "Cikarang", "Landed", "KPR",
        "3-6 mo", "2026-03-09", "2026-03-09",
        "Send SAVASA detail + price list", "Kenal dari Pak Andi. Budget fleksibel asal worth it."
    ]
    for i, val in enumerate(example, 1):
        ws.cell(row=2, column=i, value=val)

    _style_data_row(ws, 2, len(cols))

    # Color Temp cell
    temp_col = 6
    temp_cell = ws.cell(row=2, column=temp_col)
    if temp_cell.value == "HOT":
        temp_cell.fill = _fill(C_HOT)
        temp_cell.font = _font(bold=True, color="FFFFFF")
    elif temp_cell.value == "WARM":
        temp_cell.fill = _fill(C_WARM)
        temp_cell.font = _font(bold=True, color="FFFFFF")
    elif temp_cell.value == "COLD":
        temp_cell.fill = _fill(C_COLD)
        temp_cell.font = _font(bold=True, color="FFFFFF")

    # Data validations
    status_dv = DataValidation(
        type="list",
        formula1='"new,dm_sent,response,call_booked,call_done,proposal_sent,negotiation,closed_won,closed_lost,ghost"',
        allow_blank=True
    )
    temp_dv = DataValidation(
        type="list",
        formula1='"HOT,WARM,COLD"',
        allow_blank=True
    )
    source_dv = DataValidation(
        type="list",
        formula1='"IG Organic,Referral,Content,Cold Outreach,WhatsApp,Property Portal,Developer"',
        allow_blank=True
    )
    intent_dv = DataValidation(
        type="list",
        formula1='"Invest,Own use,Both"',
        allow_blank=True
    )
    type_dv = DataValidation(
        type="list",
        formula1='"Landed,Apartemen,Ruko,Kavling"',
        allow_blank=True
    )
    fin_dv = DataValidation(
        type="list",
        formula1='"Cash,KPR,Open"',
        allow_blank=True
    )
    timeline_dv = DataValidation(
        type="list",
        formula1='"ASAP,1-3 mo,3-6 mo,6-12 mo,No rush"',
        allow_blank=True
    )

    for dv, col in [
        (status_dv, 5), (temp_dv, 6), (source_dv, 4),
        (intent_dv, 7), (type_dv, 11), (fin_dv, 12), (timeline_dv, 13)
    ]:
        col_letter = get_column_letter(col)
        dv.sqref = f"{col_letter}2:{col_letter}500"
        ws.add_data_validation(dv)

    return ws


def build_revenue_sheet(wb):
    ws = wb.create_sheet("Revenue_Log")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    cols = [
        ("Date",           12),
        ("Client Name",    18),
        ("Property",       22),
        ("Project",        18),
        ("Transaction (M)", 14),
        ("Commission %",   13),
        ("Commission (M)", 14),
        ("Payment Status", 15),
        ("Notes",          28),
    ]

    for i, (header, width) in enumerate(cols, 1):
        ws.cell(row=1, column=i, value=header)
        ws.column_dimensions[get_column_letter(i)].width = width

    _style_header_row(ws, 1, len(cols))
    ws.row_dimensions[1].height = 28

    # Example closed deal
    example = [
        "2026-03-09", "Budi Santoso", "SAVASA Type 2BR Unit A12",
        "SAVASA Deltamas", 1.5, 2.5, 0.0375, "Pending", "DP sudah dibayar, menunggu akad"
    ]
    for i, val in enumerate(example, 1):
        ws.cell(row=2, column=i, value=val)
    _style_data_row(ws, 2, len(cols))

    # Commission formula (auto-calc)
    ws["G2"] = "=E2*F2/100"
    ws["G2"].number_format = '#,##0.00'

    payment_dv = DataValidation(
        type="list",
        formula1='"Pending,Partial,Received,Cancelled"',
        allow_blank=True
    )
    payment_dv.sqref = "H2:H500"
    ws.add_data_validation(payment_dv)

    return ws


def build_source_sheet(wb):
    ws = wb.create_sheet("Lead_Source")
    ws.sheet_view.showGridLines = False

    ws.cell(row=1, column=1, value="Lead Source Summary")
    ws.cell(row=1, column=1).font = _font(bold=True, size=14, color=C_ACCENT)
    ws.cell(row=1, column=1).alignment = _left()

    headers = ["Source", "Total Leads", "HOT", "WARM", "COLD", "Closed Won", "Conv. Rate"]
    widths  = [20, 13, 10, 10, 10, 13, 12]

    for i, (h, w) in enumerate(zip(headers, widths), 1):
        ws.cell(row=3, column=i, value=h)
        ws.column_dimensions[get_column_letter(i)].width = w

    _style_header_row(ws, 3, len(headers))
    ws.row_dimensions[3].height = 26

    sources = [
        "IG Organic", "Referral", "Content", "Cold Outreach",
        "WhatsApp", "Property Portal", "Developer"
    ]
    for row_i, src in enumerate(sources, 4):
        ws.cell(row=row_i, column=1, value=src)
        for col_i in range(2, 8):
            ws.cell(row=row_i, column=col_i, value=0)
        _style_data_row(ws, row_i, len(headers), alt=(row_i % 2 == 0))

    ws.row_dimensions[1].height = 30

    return ws


def main():
    if OUTPUT.exists():
        print(f"[!] File already exists: {OUTPUT}")
        confirm = input("Overwrite? (y/N): ").strip().lower()
        if confirm != "y":
            print("Aborted.")
            return

    wb = openpyxl.Workbook()
    build_pipeline_sheet(wb)
    build_revenue_sheet(wb)
    build_source_sheet(wb)

    wb.save(OUTPUT)
    print(f"[OK] RE CRM created: {OUTPUT}")
    print(f"  Sheets: Pipeline | Revenue_Log | Lead_Source")
    print(f"  Example lead pre-loaded: Budi Santoso (WARM | SAVASA)")


if __name__ == "__main__":
    main()
