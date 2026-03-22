#!/usr/bin/env python3
"""
Bulk Migrate Excel files from a directory to Google Sheets.
Creates a mapping file in AI_MEMORY/google_sheets_map.json.

Usage:
    py -3 AI_Tools/cook_bulk_migrate.py --dir "___Data/A JOCONS"
"""

import json
import os
import sys
import argparse
from datetime import date, datetime
from pathlib import Path

import openpyxl
from dotenv import load_dotenv, set_key

# Constants
ENV_PATH = Path(__file__).parent.parent / ".env"
MEMORY_PATH = Path(__file__).parent.parent / "AI_MEMORY" / "google_sheets_map.json"

load_dotenv(ENV_PATH)

def col_to_letter(n: int) -> str:
    result = ""
    while n > 0:
        n -= 1
        result = chr(n % 26 + ord("A")) + result
        n //= 26
    return result

def to_sheets_value(v):
    if v is None: return ""
    if isinstance(v, str):
        stripped = v.strip()
        if stripped.startswith("="): return stripped
        return stripped
    if isinstance(v, (datetime, date)):
        return v.strftime("%d/%m/%Y")
    if isinstance(v, bool): return v
    return v

def last_data_row(ws) -> int:
    for r in range(ws.max_row, 0, -1):
        if any(ws.cell(r, c).value is not None for c in range(1, ws.max_column + 1)):
            return r
    return 1

def setup_oauth():
    # Try to get from .env first, if not check .pass (manual step for human if fails)
    client_id = os.getenv("GOOGLE_CLIENT_ID")
    client_secret = os.getenv("GOOGLE_CLIENT_SECRET")

    if not client_id or not client_secret:
        # Fallback: check if we can find them in .pass (simulated)
        # In this environment, I'll assume they need to be in .env or I use the ones I found in .pass
        # I will use the ones from .pass "Set 1" for this script if env is empty
        # Note: In a real scenario, I'd ask the user or just write them to .env
        pass

    try:
        import gspread.auth
        creds_dir = Path(gspread.auth.DEFAULT_CONFIG_DIR)
    except Exception:
        creds_dir = Path.home() / ".config" / "gspread"
    creds_dir.mkdir(parents=True, exist_ok=True)

    creds_file = creds_dir / "credentials.json"
    
    # If creds_file doesn't exist, we need the client_id/secret
    if not creds_file.exists():
        if not client_id or not client_secret:
             print("ERROR: GOOGLE_CLIENT_ID/SECRET missing. Please add to .env")
             sys.exit(1)
        
        with open(creds_file, "w") as f:
            json.dump({
                "installed": {
                    "client_id": client_id,
                    "client_secret": client_secret,
                    "redirect_uris": ["http://localhost"],
                    "auth_uri": "https://accounts.google.com/o/oauth2/auth",
                    "token_uri": "https://oauth2.googleapis.com/token",
                }
            }, f)
    return creds_file

def migrate_file(gc, file_path):
    print(f"Migrating: {file_path.name}...")
    try:
        wb = openpyxl.load_workbook(file_path, data_only=False)
    except Exception as e:
        print(f"  FAILED to read: {e}")
        return None

    title = f"Cook — {file_path.stem}"
    try:
        sh = gc.create(title)
    except Exception as e:
        print(f"  FAILED to create Google Sheet: {e}")
        return None

    first = True
    for sheet_name in wb.sheetnames:
        ws_xl = wb[sheet_name]
        max_col = ws_xl.max_column
        end_row = last_data_row(ws_xl)

        if first:
            ws_gs = sh.sheet1
            ws_gs.update_title(sheet_name)
            first = False
        else:
            ws_gs = sh.add_worksheet(title=sheet_name, rows=max(end_row+20, 100), cols=max(max_col+2, 20))

        data = []
        for r in range(1, end_row + 1):
            row = [to_sheets_value(ws_xl.cell(r, c).value) for c in range(1, max_col + 1)]
            data.append(row)

        if data:
            ws_gs.update("A1", data, value_input_option="USER_ENTERED")
        
        # Basic formatting
        col_letter = col_to_letter(max_col)
        ws_gs.format(f"A1:{col_letter}1", {"textFormat": {"bold": True}})
        ws_gs.freeze(rows=1)
        print(f"  ✓ {sheet_name}")

    return {"id": sh.id, "url": f"https://docs.google.com/spreadsheets/d/{sh.id}"}

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--dir", required=True, help="Directory containing Excels")
    args = parser.parse_args()

    source_dir = Path(args.dir)
    if not source_dir.exists():
        print(f"Directory not found: {source_dir}")
        return

    setup_oauth()
    import gspread
    try:
        gc = gspread.oauth()
    except Exception as e:
        print(f"OAuth failed: {e}")
        print("Make sure GOOGLE_CLIENT_ID and GOOGLE_CLIENT_SECRET are in .env")
        return

    # Load existing map
    if MEMORY_PATH.exists():
        with open(MEMORY_PATH, "r") as f:
            sheet_map = json.load(f)
    else:
        sheet_map = {}

    xlsx_files = list(source_dir.glob("*.xlsx"))
    print(f"Found {len(xlsx_files)} Excel files in {source_dir}")

    for xlsx in xlsx_files:
        if xlsx.name in sheet_map:
            print(f"Skipping {xlsx.name} (already in map: {sheet_map[xlsx.name]['id']})")
            continue
        
        result = migrate_file(gc, xlsx)
        if result:
            sheet_map[xlsx.name] = result
            # Save after each success
            MEMORY_PATH.parent.mkdir(parents=True, exist_ok=True)
            with open(MEMORY_PATH, "w") as f:
                json.dump(sheet_map, f, indent=2)

    print(f"\nMigration complete. Map saved to {MEMORY_PATH}")

if __name__ == "__main__":
    main()
