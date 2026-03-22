"""
Setter & Cash Dashboard — Excel Infrastructure
Creates a multi-sheet Excel workbook:
  1. CRM         — Lead tracker (data entry)
  2. EOD         — Daily setter KPI input
  3. SetterDash  — Setter performance dashboard (formulas)
  4. CashDash    — Sales/cash dashboard (formulas)
"""

import os
from datetime import datetime, date
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import DataBarRule

# ── Styling constants ──────────────────────────────────────────────
BLACK = "1A1A2E"
DARK = "16213E"
CARD_BG = "0F3460"
ACCENT = "E94560"
WHITE = "FFFFFF"
GRAY = "B0B0B0"
GREEN = "00C853"
YELLOW = "FFD600"
RED = "FF1744"
LIGHT_DARK = "222244"

font_header = Font(name="Aptos", bold=True, color=WHITE, size=11)
font_subheader = Font(name="Aptos", bold=True, color=GRAY, size=9)
font_data = Font(name="Aptos", color=WHITE, size=10)
font_kpi_value = Font(name="Aptos", bold=True, color=WHITE, size=22)
font_kpi_label = Font(name="Aptos", color=GRAY, size=9)
font_kpi_sub = Font(name="Aptos", color=ACCENT, size=10)

fill_black = PatternFill("solid", fgColor=BLACK)
fill_dark = PatternFill("solid", fgColor=DARK)
fill_card = PatternFill("solid", fgColor=CARD_BG)
fill_accent = PatternFill("solid", fgColor=ACCENT)
fill_header = PatternFill("solid", fgColor="0D1B2A")
fill_row_alt = PatternFill("solid", fgColor="1B2838")
fill_green = PatternFill("solid", fgColor="1B5E20")
fill_red = PatternFill("solid", fgColor="B71C1C")

thin_border = Border(
    left=Side(style="thin", color="333355"),
    right=Side(style="thin", color="333355"),
    top=Side(style="thin", color="333355"),
    bottom=Side(style="thin", color="333355"),
)

align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)


def style_sheet_dark(ws):
    """Apply dark background to entire sheet."""
    ws.sheet_properties.tabColor = ACCENT
    for row in ws.iter_rows(min_row=1, max_row=200, max_col=40):
        for cell in row:
            cell.fill = fill_black
            cell.font = font_data
            cell.alignment = align_center
            cell.border = thin_border


def make_kpi_card(ws, row, col, label, formula, fmt="0", width=2):
    """Create a KPI card spanning `width` columns."""
    # Label
    cell_label = ws.cell(row=row, column=col, value=label)
    cell_label.font = font_kpi_label
    cell_label.fill = fill_card
    cell_label.alignment = align_center
    if width > 1:
        ws.merge_cells(
            start_row=row, start_column=col,
            end_row=row, end_column=col + width - 1
        )
    # Value
    cell_val = ws.cell(row=row + 1, column=col, value=formula)
    cell_val.font = font_kpi_value
    cell_val.fill = fill_card
    cell_val.alignment = align_center
    cell_val.number_format = fmt
    if width > 1:
        ws.merge_cells(
            start_row=row + 1, start_column=col,
            end_row=row + 1, end_column=col + width - 1
        )
    # Style remaining cells in card
    for r in [row, row + 1]:
        for c in range(col, col + width):
            ws.cell(row=r, column=c).fill = fill_card
            ws.cell(row=r, column=c).border = thin_border


# ════════════════════════════════════════════════════════════════════
# SHEET 1: CRM
# ════════════════════════════════════════════════════════════════════
def build_crm(wb):
    ws = wb.active
    ws.title = "CRM"
    style_sheet_dark(ws)

    headers = [
        "LeadAge\n(Days)", "24-Convo", "24Rate", "Username", "LeadType",
        "Status", "Mindmap", "Objections", "Objection Type", "Booked",
        "Show", "Close", "Close Rate", "LeadDate", "Close Date",
        "Lead To Close", "24ConvoDate", "Date", "BookName", "CloseName",
        "Package", "CC", "LTV", "Commission", "WA Number\n(Freebie)",
        "Email\n(Freebie)", "LeadSource", "Content Feedback", "Proposed",
        "Live Calls Date", "Show Up Rate"
    ]

    col_widths = [
        10, 9, 9, 14, 13, 12, 20, 20, 14, 8,
        8, 10, 10, 10, 10, 12, 12, 10, 12, 12,
        14, 14, 12, 12, 18, 22, 12, 18, 9, 14, 12
    ]

    # Header row
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(i)].width = col_widths[i - 1]

    # Sample data
    sample_data = [
        [8, False, "33.3%", "Sapri", "Lead", "Unqualified", "", "", "",
         False, False, "Not Started", "100.0%", "09/02", "20/02",
         "11 days", "09/02", "20/01", "Jepri", "Supri", "", "", "", "",
         "", "", "Cold", "", False, "20/01", "33.33%"],
        [14, False, "33.3%", "Jule", "ICP", "Unqualified", "", "", "",
         False, False, "Not Started", "100.0%", "03/02", "20/02",
         "17 days", "09/02", "20/01", "Epstein", "Supri", "", "", "", "",
         "", "", "Cold", "", False, "20/01", "33.33%"],
        [13, False, "33.3%", "Bigmo", "Follower", "Unqualified", "", "", "",
         False, False, "Not Started", "100.0%", "04/02", "20/02",
         "16 days", "08/02", "20/01", "Epstein", "Supri", "", "", "", "",
         "", "", "Cold", "", False, "20/01", "33.33%"],
        [13, True, "33.3%", "astuti", "ICPLongForm", "Qualified", "", "", "",
         True, False, "Not Started", "100.0%", "04/02", "20/02",
         "16 days", "08/02", "20/01", "Epstein", "Supri", "", "", "", "",
         "", "", "Cold", "", False, "20/01", "33.33%"],
        [13, True, "33.3%", "jerome", "ICP-Freebie", "Unqualified",
         "ambil free content pack\nblm ada uang", "", "",
         True, True, "Closed", "100.0%", "04/02", "20/02",
         "16 days", "08/02", "20/01", "Epstein", "Supri", "", "", "", "",
         "", "", "Cold", "", False, "20/01", "33.33%"],
        [13, False, "33.3%", "boris", "ICP-CTA", "Qualified", "1", "", "",
         True, True, "Closed", "100.0%", "04/02", "20/02",
         "16 days", "08/02", "20/01", "Epstein", "Supri",
         "1-1 Coaching", "", "", "", "", "", "Cold", "", False, "20/01", "33.33%"],
    ]

    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = font_data
            cell.fill = fill_row_alt if row_idx % 2 == 0 else fill_black
            cell.alignment = align_center
            cell.border = thin_border

    # Conditional: color Status column
    for row_idx in range(2, 8):
        status_cell = ws.cell(row=row_idx, column=6)
        if status_cell.value == "Qualified":
            status_cell.fill = fill_green
        elif status_cell.value == "Unqualified":
            status_cell.fill = fill_red

        close_cell = ws.cell(row=row_idx, column=12)
        if close_cell.value == "Closed":
            close_cell.fill = fill_green
            close_cell.font = Font(name="Aptos", bold=True, color=WHITE, size=10)

    # Data validation hints
    ws.cell(row=1, column=33, value="LEAD TYPES →").font = font_subheader
    lead_types = ["Lead", "ICP", "Follower", "ICPLongForm", "ICP-Freebie", "ICP-CTA"]
    for i, lt in enumerate(lead_types):
        ws.cell(row=2 + i, column=33, value=lt).font = font_data

    ws.cell(row=1, column=34, value="STATUS →").font = font_subheader
    statuses = ["Unqualified", "Qualified", "Not Started", "Closed", "No Show", "Nurturing"]
    for i, s in enumerate(statuses):
        ws.cell(row=2 + i, column=34, value=s).font = font_data

    ws.cell(row=1, column=35, value="SOURCES →").font = font_subheader
    sources = ["Cold", "Warm", "Inbound", "Referral", "Ads", "CTA-Keyword", "TikTok"]
    for i, s in enumerate(sources):
        ws.cell(row=2 + i, column=35, value=s).font = font_data

    ws.freeze_panes = "E2"
    ws.auto_filter.ref = f"A1:AE1"
    return ws


# ════════════════════════════════════════════════════════════════════
# SHEET 2: EOD (Setter Daily Report)
# ════════════════════════════════════════════════════════════════════
def build_eod(wb):
    ws = wb.create_sheet("EOD")
    style_sheet_dark(ws)

    headers = [
        "Date", "Setter Name", "Productivity\n(1-10)", "Hours Worked",
        "Inbox Checked", "New Convo's", "Follow Ups Sent",
        "YT/LongForm Sent", "Calls Proposed", "Calls Booked (DM)",
        "Qualified Bookings", "Unqualified Bookings", "ICP Leads Today",
        "Convos Nurtured", "Calls on Calendar", "Calls Showed",
        "Sets Closed", "What Went Well", "What Could Be Better",
        "Situation Update", "Top Objection Today", "Content Feedback"
    ]

    col_widths = [
        12, 14, 12, 12, 12, 12, 14,
        14, 13, 15, 15, 16, 13,
        14, 14, 12, 11, 28, 28,
        28, 24, 24
    ]

    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(i)].width = col_widths[i - 1]

    # Sample EOD entries
    sample_eod = [
        ["17/02/2026", "Nick", 10, 8, 100, 2, 40, 7, 4, 5, 5, 0, 2, 30, 1, 1, 0,
         "5 qualified calls booked, good call with team and Josh on content",
         "More volume, more calls on calendar, message follow-ups from YT",
         "More leads coming in from reels", "I don't have money", "Need ROI content"],
        ["16/02/2026", "Nick", 8, 7, 85, 3, 35, 5, 3, 4, 3, 1, 3, 25, 2, 1, 0,
         "Good nurture flow, 3 ICP found", "Need more cold scraping",
         "Inbox getting busy", "Need to think about it", "Comparison reel would help"],
    ]

    for row_idx, row_data in enumerate(sample_eod, 2):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = font_data
            cell.fill = fill_row_alt if row_idx % 2 == 0 else fill_black
            cell.alignment = align_center if col_idx <= 17 else align_left
            cell.border = thin_border

    # Targets row (row 50 — reference)
    ws.cell(row=50, column=1, value="TARGETS").font = font_header
    ws.cell(row=50, column=1).fill = fill_accent
    targets = [
        "", "", 8, 7, 100, 2, 40, 7, 4, 5, 5, 0, 2, 30, 1, 1, 0
    ]
    for i, t in enumerate(targets, 1):
        if t:
            ws.cell(row=50, column=i, value=t).font = font_header
            ws.cell(row=50, column=i).fill = fill_accent

    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:V1"
    return ws


# ════════════════════════════════════════════════════════════════════
# SHEET 3: Setter Dashboard
# ════════════════════════════════════════════════════════════════════
def build_setter_dash(wb):
    ws = wb.create_sheet("SetterDash")
    style_sheet_dark(ws)

    # Title
    ws.merge_cells("A1:N1")
    title = ws.cell(row=1, column=1, value="SETTER PERFORMANCE DASHBOARD")
    title.font = Font(name="Aptos", bold=True, color=WHITE, size=16)
    title.fill = fill_dark
    title.alignment = align_center

    ws.row_dimensions[1].height = 35

    # ── Row 3-4: KPI Cards ──
    # QC Booked (from CRM: count where Booked=TRUE AND Status=Qualified)
    make_kpi_card(ws, 3, 1, "Q.C Booked",
                  '=COUNTIFS(CRM!J:J,TRUE,CRM!F:F,"Qualified")', "0", 2)

    # Convo → Qualified %
    make_kpi_card(ws, 3, 3, "Convo - Qualified %",
                  '=IFERROR(COUNTIFS(CRM!F:F,"Qualified")/COUNTA(CRM!D2:D1000),0)', "0.0%", 2)

    # Commission (sum from CRM)
    make_kpi_card(ws, 3, 5, "Commission",
                  '=SUM(CRM!X2:X1000)', "#,##0", 2)

    # Show Up Rate
    make_kpi_card(ws, 3, 7, "Show Up Rate",
                  '=IFERROR(COUNTIFS(CRM!K:K,TRUE)/COUNTIFS(CRM!J:J,TRUE),0)', "0%", 2)

    # ── Row 6-7: Secondary KPIs ──
    # Avg ICP / day (from EOD)
    make_kpi_card(ws, 6, 1, "Avg ICP / Day",
                  '=IFERROR(AVERAGE(EOD!M2:M1000),0)', "0", 2)

    # Avg Follow Ups / day
    make_kpi_card(ws, 6, 3, "Avg Follow Ups",
                  '=IFERROR(AVERAGE(EOD!G2:G1000),0)', "0", 2)

    # Avg QC Booked / day
    make_kpi_card(ws, 6, 5, "Avg Q.C Booked",
                  '=IFERROR(AVERAGE(EOD!K2:K1000),0)', "0", 2)

    # Avg Productivity Score
    make_kpi_card(ws, 6, 7, "Avg Productivity",
                  '=IFERROR(AVERAGE(EOD!C2:C1000),0)', "0.0", 2)

    # ── Row 9: Section header ──
    ws.merge_cells("A9:H9")
    ws.cell(row=9, column=1, value="DAILY EOD METRICS").font = font_header
    ws.cell(row=9, column=1).fill = fill_header

    # ── Row 10+: Pull EOD data as a mirror ──
    eod_cols = [
        "Date", "Setter", "Score", "Hours", "Inbox",
        "New Convo", "Follow Ups", "QC Booked"
    ]
    for i, h in enumerate(eod_cols, 1):
        cell = ws.cell(row=10, column=i, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center

    # Link to EOD data (first 30 rows)
    for r in range(11, 41):
        eod_row = r - 9  # maps to EOD row 2, 3, ...
        ws.cell(row=r, column=1, value=f"=EOD!A{eod_row}").font = font_data
        ws.cell(row=r, column=2, value=f"=EOD!B{eod_row}").font = font_data
        ws.cell(row=r, column=3, value=f"=EOD!C{eod_row}").font = font_data
        ws.cell(row=r, column=4, value=f"=EOD!D{eod_row}").font = font_data
        ws.cell(row=r, column=5, value=f"=EOD!E{eod_row}").font = font_data
        ws.cell(row=r, column=6, value=f"=EOD!F{eod_row}").font = font_data
        ws.cell(row=r, column=7, value=f"=EOD!G{eod_row}").font = font_data
        ws.cell(row=r, column=8, value=f"=EOD!K{eod_row}").font = font_data
        for c in range(1, 9):
            ws.cell(row=r, column=c).fill = fill_row_alt if r % 2 == 0 else fill_black
            ws.cell(row=r, column=c).alignment = align_center
            ws.cell(row=r, column=c).border = thin_border

    # ── Qualified Bookings Chart ──
    chart = BarChart()
    chart.type = "col"
    chart.title = "Qualified Bookings Per Day"
    chart.y_axis.title = "# Bookings"
    chart.x_axis.title = "Date"
    chart.style = 10
    chart.width = 22
    chart.height = 12

    data_ref = Reference(ws, min_col=8, min_row=10, max_row=40)
    cats_ref = Reference(ws, min_col=1, min_row=11, max_row=40)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.shape = 4
    ws.add_chart(chart, "J3")

    # Column widths
    for c in range(1, 15):
        ws.column_dimensions[get_column_letter(c)].width = 14

    return ws


# ════════════════════════════════════════════════════════════════════
# SHEET 4: Cash Dashboard
# ════════════════════════════════════════════════════════════════════
def build_cash_dash(wb):
    ws = wb.create_sheet("CashDash")
    style_sheet_dark(ws)

    # Title
    ws.merge_cells("A1:N1")
    title = ws.cell(row=1, column=1, value="CASH COLLECTED DASHBOARD")
    title.font = Font(name="Aptos", bold=True, color=WHITE, size=16)
    title.fill = fill_dark
    title.alignment = align_center
    ws.row_dimensions[1].height = 35

    # ── Row 3-4: Top KPI Row ──
    # QC's Booked
    make_kpi_card(ws, 3, 1, "QC's Booked",
                  '=COUNTIFS(CRM!J:J,TRUE,CRM!F:F,"Qualified")', "0", 2)

    # Close Rate
    make_kpi_card(ws, 3, 3, "Close Rate",
                  '=IFERROR(COUNTIFS(CRM!L:L,"Closed")/COUNTIFS(CRM!K:K,TRUE),0)',
                  "0%", 2)

    # Show Up Rate
    make_kpi_card(ws, 3, 5, "Show Up Rate",
                  '=IFERROR(COUNTIFS(CRM!K:K,TRUE)/COUNTIFS(CRM!J:J,TRUE),0)',
                  "0%", 2)

    # New Cash (CC from leads where LeadSource != referral-ish)
    make_kpi_card(ws, 3, 7, "New Cash",
                  '=SUMPRODUCT((CRM!L2:L1000="Closed")*(CRM!V2:V1000))',
                  '#,##0', 2)

    # Cash Collected (total)
    make_kpi_card(ws, 3, 9, "Cash Collected",
                  '=SUM(CRM!V2:V1000)', '#,##0', 2)

    # ── Row 6-7: Secondary KPIs ──
    # Existing Amount (LTV)
    make_kpi_card(ws, 6, 1, "Existing Amount (LTV)",
                  '=SUM(CRM!W2:W1000)', '#,##0', 2)

    # New : Existing ratio
    make_kpi_card(ws, 6, 3, "New$ : Existing$",
                  '=IFERROR(SUMPRODUCT((CRM!L2:L1000="Closed")*(CRM!V2:V1000))/SUM(CRM!W2:W1000),"N/A")',
                  '0.0 ": 1"', 2)

    # Avg CC
    make_kpi_card(ws, 6, 5, "Avg. C.C",
                  '=IFERROR(AVERAGE(IF(CRM!V2:V1000>0,CRM!V2:V1000)),0)',
                  '#,##0', 2)

    # Total Commission
    make_kpi_card(ws, 6, 7, "Total Commission",
                  '=SUM(CRM!X2:X1000)', '#,##0', 2)

    # Payment Plans Due
    make_kpi_card(ws, 6, 9, "Payment Plans Due",
                  '=0', '#,##0', 2)  # Placeholder — needs payment plan sheet

    # ── Row 9: Recent Closes ──
    ws.merge_cells("A9:F9")
    ws.cell(row=9, column=1, value="RECENT CLOSES").font = font_header
    ws.cell(row=9, column=1).fill = fill_header

    close_headers = ["#", "Date", "Client", "Package", "Cash Collected", "Closer"]
    for i, h in enumerate(close_headers, 1):
        cell = ws.cell(row=10, column=i, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border

    # Pull recent closes from CRM (manual — formula references)
    # These show the most recent rows where Close = "Closed"
    for r in range(11, 31):
        ws.cell(row=r, column=1, value=r - 10).font = font_data
        for c in range(1, 7):
            ws.cell(row=r, column=c).fill = fill_row_alt if r % 2 == 0 else fill_black
            ws.cell(row=r, column=c).alignment = align_center
            ws.cell(row=r, column=c).border = thin_border

    # ── Row 9 col H: Payment Plans ──
    ws.merge_cells("H9:L9")
    ws.cell(row=9, column=8, value="PAYMENT PLANS").font = font_header
    ws.cell(row=9, column=8).fill = fill_header

    pp_headers = ["#", "Due Date", "Client", "Status", "Amount Due"]
    for i, h in enumerate(pp_headers):
        cell = ws.cell(row=10, column=8 + i, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border

    for r in range(11, 31):
        for c in range(8, 13):
            ws.cell(row=r, column=c).fill = fill_row_alt if r % 2 == 0 else fill_black
            ws.cell(row=r, column=c).alignment = align_center
            ws.cell(row=r, column=c).border = thin_border

    # ── Cash Collected bar chart ──
    # We'll create a mini summary area for chart data
    ws.cell(row=33, column=1, value="MONTHLY SUMMARY").font = font_header
    ws.cell(row=33, column=1).fill = fill_header
    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
              "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    for i, m in enumerate(months):
        ws.cell(row=34, column=1 + i, value=m).font = font_subheader
        ws.cell(row=35, column=1 + i, value=0).font = font_data  # Cash collected per month
        ws.cell(row=35, column=1 + i).number_format = '#,##0'

    chart = BarChart()
    chart.type = "col"
    chart.title = "Cash Collected by Month"
    chart.y_axis.title = "Amount"
    chart.style = 10
    chart.width = 28
    chart.height = 12

    data_ref = Reference(ws, min_col=1, max_col=12, min_row=35)
    cats_ref = Reference(ws, min_col=1, max_col=12, min_row=34)
    chart.add_data(data_ref, from_rows=True)
    chart.set_categories(cats_ref)
    ws.add_chart(chart, "A37")

    # Column widths
    for c in range(1, 15):
        ws.column_dimensions[get_column_letter(c)].width = 15

    return ws


# ════════════════════════════════════════════════════════════════════
# SHEET 5: PaymentPlans
# ════════════════════════════════════════════════════════════════════
def build_payment_plans(wb):
    ws = wb.create_sheet("PaymentPlans")
    style_sheet_dark(ws)

    headers = [
        "#", "Payment Due Date", "Client Name", "Package",
        "Status", "Payment Due", "Amount Paid", "Remaining",
        "Notes"
    ]
    col_widths = [5, 16, 18, 16, 10, 14, 14, 14, 24]

    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(i)].width = col_widths[i - 1]

    # Sample payment plan data
    sample = [
        [1, "01/03/2026", "Zain Bilal", "Bedah Digital", "paid", 1875, 1875, 0, ""],
        [2, "01/03/2026", "Nathan Niu", "Bedah Digital", "paid", 4000, 4000, 0, ""],
        [3, "01/03/2026", "Yash Narendran", "AI & Scale", "not paid", 4000, 0, 4000, "Follow up sent"],
        [4, "01/03/2026", "Rolando Balcarcel", "Bedah Digital", "not paid", 4000, 0, 4000, ""],
        [5, "01/03/2026", "Damien Law", "AI & Scale", "not paid", 10000, 0, 10000, ""],
    ]

    for row_idx, row_data in enumerate(sample, 2):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = font_data
            cell.fill = fill_row_alt if row_idx % 2 == 0 else fill_black
            cell.alignment = align_center if col_idx <= 8 else align_left
            cell.border = thin_border

        # Color status
        status_cell = ws.cell(row=row_idx, column=5)
        if status_cell.value == "paid":
            status_cell.fill = fill_green
        elif status_cell.value == "not paid":
            status_cell.fill = fill_red

    # Remaining formula
    for r in range(2, 7):
        ws.cell(row=r, column=8, value=f"=F{r}-G{r}")
        ws.cell(row=r, column=8).number_format = '#,##0'

    # Summary
    ws.cell(row=9, column=5, value="TOTAL DUE:").font = font_header
    ws.cell(row=9, column=6, value="=SUM(F2:F1000)").font = font_kpi_value
    ws.cell(row=9, column=6).font = Font(name="Aptos", bold=True, color=ACCENT, size=14)
    ws.cell(row=9, column=6).number_format = '#,##0'

    ws.cell(row=10, column=5, value="TOTAL PAID:").font = font_header
    ws.cell(row=10, column=6, value="=SUM(G2:G1000)").font = font_kpi_value
    ws.cell(row=10, column=6).font = Font(name="Aptos", bold=True, color=GREEN, size=14)
    ws.cell(row=10, column=6).number_format = '#,##0'

    ws.cell(row=11, column=5, value="OUTSTANDING:").font = font_header
    ws.cell(row=11, column=6, value="=SUM(H2:H1000)").font = font_kpi_value
    ws.cell(row=11, column=6).font = Font(name="Aptos", bold=True, color=RED, size=14)
    ws.cell(row=11, column=6).number_format = '#,##0'

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:I1"
    return ws


# ════════════════════════════════════════════════════════════════════
# MAIN
# ════════════════════════════════════════════════════════════════════
def main():
    wb = Workbook()

    print("Building CRM sheet...")
    build_crm(wb)

    print("Building EOD sheet...")
    build_eod(wb)

    print("Building Setter Dashboard...")
    build_setter_dash(wb)

    print("Building Cash Dashboard...")
    build_cash_dash(wb)

    print("Building Payment Plans...")
    build_payment_plans(wb)

    # Save
    output_dir = os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        ".tmp"
    )
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, "JO_Consult_Dashboard.xlsx")
    wb.save(output_path)
    print(f"\nDashboard saved to: {output_path}")
    print("Sheets: CRM | EOD | SetterDash | CashDash | PaymentPlans")


if __name__ == "__main__":
    main()
