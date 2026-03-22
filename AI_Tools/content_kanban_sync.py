"""
Content Kanban → Excel Performance Tracker Sync Tool

Reads Obsidian Kanban board (VLT_Content/01 HMN_Command/CONTENT SYSTEM.md),
extracts cards from "Scheduled" lane, syncs to Excel performance tracker.

Usage:
    py -3 AI_Tools/content_kanban_sync.py
"""

import re
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


def parse_kanban(kanban_text: str) -> list[dict]:
    """
    Parse Obsidian Kanban markdown. Extract cards from 'Scheduled' lane only.

    Returns list of dicts with keys: title, platform, post_date, captions
    """
    # Split by lane headers
    lanes = re.split(r'^## ', kanban_text, flags=re.MULTILINE)

    items = []
    for lane in lanes:
        if not lane.startswith('Scheduled'):
            continue

        # Extract each card block (starts with "- [ ] ###")
        card_blocks = re.findall(
            r'- \[ \] ### \[([^\]]+)\] ([^\n]+)\n((?:\t[^\n]+\n)*)',
            lane,
            re.MULTILINE
        )

        for tag, title, metadata_block in card_blocks:
            # Parse metadata
            platform = extract_metadata(metadata_block, 'Platform')
            post_date = extract_metadata(metadata_block, 'Post Date')
            captions = extract_metadata(metadata_block, 'Captions')

            # Clean platform (strip emoji prefix and standardize)
            platform = clean_platform(platform)

            items.append({
                'title': title.strip(),
                'platform': platform,
                'post_date': post_date if post_date != 'TBD' else '',
                'captions': captions if captions != 'TBD' else '',
            })

    return items


def extract_metadata(metadata_block: str, key: str) -> str:
    """Extract value from metadata block, e.g., '**Platform:** 📸 Instagram'"""
    pattern = rf'\*\*{key}:\*\*\s*(.+?)(?:\n|$)'
    match = re.search(pattern, metadata_block)
    return match.group(1).strip() if match else ''


def clean_platform(platform_str: str) -> str:
    """
    Normalize platform string. Remove emoji, standardize names.
    Example: '📸 Instagram / 📱 TikTok' -> 'IG Reel'
    """
    if not platform_str:
        return ''

    # Remove emoji first
    platform_clean = re.sub(r'[📱📸🎥📊]', '', platform_str).strip()

    # Split by "/" and process first platform only
    first_platform = platform_clean.split('/')[0].split(',')[0].strip().lower()

    # Map variants to standard names (order matters - check specific patterns first)
    mapping = [
        (r'instagram.*story|ig.*story', 'IG Story'),
        (r'youtube.*short|yt.*short', 'YouTube Short'),
        (r'youtube.*long|yt.*long', 'YouTube Long'),
        (r'instagram|ig\s*reel|ig\s*reels', 'IG Reel'),
        (r'tiktok', 'TikTok'),
        (r'threads', 'Threads'),
        (r'analysis', 'Analysis'),
    ]

    for pattern, standard_name in mapping:
        if re.search(pattern, first_platform):
            return standard_name

    return first_platform.title() if first_platform else ''


def load_or_create_tracker(filepath: Path) -> tuple:
    """
    Load existing Excel or create new one with structure.
    Returns (workbook, data_list, sync_log_list)
    """
    if filepath.exists():
        from openpyxl import load_workbook
        wb = load_workbook(filepath)
        log_sheet = wb['Log']
        data = []

        # Load existing data (skip header row 1)
        for row in log_sheet.iter_rows(min_row=2, values_only=False):
            if row[0].value is None:  # Stop at first empty row
                break
            data.append(row)

        sync_sheet = wb['Sync Log']
        sync_log = []
        for row in sync_sheet.iter_rows(min_row=2, values_only=False):
            if row[0].value is None:
                break
            sync_log.append(row)

        return wb, data, sync_log
    else:
        wb = Workbook()
        wb.remove(wb.active)

        # Create Log sheet
        log_sheet = wb.create_sheet('Log')
        headers = [
            '#', 'Content Title', 'Platform', 'Scheduled Date', 'Post Date', 'Status',
            'Views 24h', 'Views 7d', 'Views 30d', 'Reach', 'Saves', 'Shares', 'Comments',
            'Likes', 'Watch Time %', 'Hook Rate %', 'Completion %', 'Bio Visits', 'DMs',
            'Leads', 'Follower Gain', 'CTR Score (1–10)', 'Notes'
        ]
        log_sheet.append(headers)

        # Create Summary sheet
        summary_sheet = wb.create_sheet('Summary')
        summary_sheet.append(['Metric', 'Value'])

        # Create Sync Log sheet
        sync_sheet = wb.create_sheet('Sync Log')
        sync_sheet.append(['Timestamp', 'Items Found', 'New Rows Added', 'Skipped (Duplicate)'])

        return wb, [], []


def get_existing_titles(data: list) -> set:
    """Extract all existing content titles from loaded data."""
    return {row[1].value for row in data if row[1].value}


def add_new_rows(log_sheet, data: list, kanban_items: list) -> int:
    """
    Add new kanban items to Excel. Dedup by title.
    Returns count of new rows added.
    """
    existing_titles = get_existing_titles(data)
    new_count = 0

    for item in kanban_items:
        if item['title'] in existing_titles:
            continue

        # Determine row number (after existing data + header)
        row_num = len(data) + 2

        # Add row
        log_sheet.cell(row_num, 1).value = row_num - 1  # Auto-number
        log_sheet.cell(row_num, 2).value = item['title']
        log_sheet.cell(row_num, 3).value = item['platform']
        log_sheet.cell(row_num, 4).value = datetime.now().strftime('%d-%b-%y')
        log_sheet.cell(row_num, 5).value = item['post_date']
        log_sheet.cell(row_num, 6).value = 'Scheduled'

        new_count += 1

    return new_count


def format_excel(workbook):
    """Apply formatting: headers, dropdowns, color scales, formulas."""
    log_sheet = workbook['Log']
    summary_sheet = workbook['Summary']

    # === LOG SHEET FORMATTING ===

    # Freeze header row
    log_sheet.freeze_panes = 'A2'

    # Header row styling
    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for cell in log_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    # Column widths
    widths = {
        'A': 4, 'B': 20, 'C': 15, 'D': 14, 'E': 12, 'F': 12,
        'G': 12, 'H': 12, 'I': 12, 'J': 12, 'K': 10, 'L': 10, 'M': 12,
        'N': 10, 'O': 13, 'P': 12, 'Q': 13, 'R': 12, 'S': 10, 'T': 10,
        'U': 12, 'V': 12, 'W': 20
    }
    for col, width in widths.items():
        log_sheet.column_dimensions[col].width = width

    # Data row formatting (find last row with data)
    last_row = log_sheet.max_row
    for row_num in range(2, last_row + 1):
        # Alternating row fill
        if row_num % 2 == 0:
            row_fill = PatternFill(start_color='E8F0F8', end_color='E8F0F8', fill_type='solid')
        else:
            row_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

        for cell in log_sheet[row_num]:
            cell.fill = row_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left', vertical='center')

            # Number formatting
            if cell.column in [1, 7, 8, 9, 10, 11, 12, 13, 14, 18, 19, 20]:
                cell.number_format = '#,##0'
            elif cell.column in [4, 5, 15, 16, 17]:
                cell.number_format = 'dd-mmm-yy'

    # Status dropdown
    status_dv = DataValidation('list', formula1='"Scheduled,Posted,Analyzing,Archived"')
    status_dv.add('F2:F1000')
    log_sheet.add_data_validation(status_dv)

    # Platform dropdown
    platform_dv = DataValidation('list', formula1='"IG Reel,IG Story,TikTok,YouTube Short,Threads,YouTube Long"')
    platform_dv.add('C2:C1000')
    log_sheet.add_data_validation(platform_dv)

    # CTR Score color scale (column V, 1–10)
    from openpyxl.formatting.rule import ColorScaleRule
    red_to_green = ColorScaleRule(
        start_type='num', start_value=1, start_color='F8696B',
        mid_type='num', mid_value=5, mid_color='FFEB84',
        end_type='num', end_value=10, end_color='63BE7B'
    )
    log_sheet.conditional_formatting.add(f'V2:V{last_row}', red_to_green)

    # === SUMMARY SHEET FORMATTING ===

    summary_sheet.column_dimensions['A'].width = 30
    summary_sheet.column_dimensions['B'].width = 15

    # Summary header
    for cell in summary_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border


def append_sync_log(workbook, found_count: int, added_count: int, skipped_count: int):
    """Append entry to Sync Log sheet."""
    sync_sheet = workbook['Sync Log']
    row_num = sync_sheet.max_row + 1

    sync_sheet.cell(row_num, 1).value = datetime.now().isoformat()
    sync_sheet.cell(row_num, 2).value = found_count
    sync_sheet.cell(row_num, 3).value = added_count
    sync_sheet.cell(row_num, 4).value = skipped_count


def main():
    """Main sync workflow."""
    # Paths
    kanban_path = Path(
        'VLT_Content/01 HMN_Command/CONTENT SYSTEM.md'
    )
    tracker_dir = Path('01 HMN__Command/00_CONTENT')
    tracker_path = tracker_dir / 'Content_Performance_Tracker.xlsx'

    # Create directory if needed
    tracker_dir.mkdir(parents=True, exist_ok=True)

    # Read Kanban
    if not kanban_path.exists():
        print(f"[ERROR] Kanban file not found: {kanban_path}")
        return

    kanban_text = kanban_path.read_text(encoding='utf-8')
    kanban_items = parse_kanban(kanban_text)

    if not kanban_items:
        print("[INFO] No items found in 'Scheduled' lane")
        return

    # Load or create tracker
    workbook, existing_data, _ = load_or_create_tracker(tracker_path)
    log_sheet = workbook['Log']

    existing_titles = get_existing_titles(existing_data)
    skipped = sum(1 for item in kanban_items if item['title'] in existing_titles)
    added = add_new_rows(log_sheet, existing_data, kanban_items)

    # Format and save
    format_excel(workbook)
    append_sync_log(workbook, len(kanban_items), added, skipped)
    workbook.save(tracker_path)

    # Summary
    print("[SUCCESS] Sync complete!")
    print(f"   Found in 'Scheduled': {len(kanban_items)}")
    print(f"   New rows added: {added}")
    print(f"   Skipped (duplicate): {skipped}")
    print(f"   Saved to: {tracker_path}")


if __name__ == '__main__':
    main()
