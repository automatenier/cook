import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC = r"01 HMN__Command/00_CONTENT/Authority_Proof_Planner.csv"
DST = r"01 HMN__Command/00_CONTENT/Authority_Proof_Planner.xlsx"

# Colors
HEADER_BG   = "1A1A2E"   # deep navy
HEADER_FG   = "FFFFFF"
ROW_ODD     = "F4F6FB"
ROW_EVEN    = "FFFFFF"
ACCENT      = "4F46E5"   # indigo for ID column
ACCENT_FG   = "FFFFFF"

thin_side = Side(style="thin", color="D1D5DB")
border    = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Authority Proof Planner"

# Column widths (A=ID, B=Visual_Type, C=Environment, D=Hook, E=Mechanism, F=Proof, G=Variation_Count)
col_widths = [5, 28, 22, 48, 48, 48, 12, 40]

with open(SRC, newline="", encoding="utf-8") as f:
    reader = list(csv.reader(f))

headers = reader[0]
rows    = [r for r in reader[1:] if any(c.strip() for c in r)]

# --- Header row ---
for col_idx, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col_idx, value=header.replace("_", " "))
    cell.font      = Font(bold=True, color=HEADER_FG, size=11, name="Calibri")
    cell.fill      = PatternFill("solid", fgColor=HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = border

ws.row_dimensions[1].height = 30

# --- Data rows ---
for row_idx, row in enumerate(rows, start=2):
    is_odd = (row_idx % 2 == 0)
    bg = ROW_ODD if is_odd else ROW_EVEN

    for col_idx, value in enumerate(row, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value.strip())
        cell.border    = border

        # ID column gets accent colour
        if col_idx == 1:
            cell.fill      = PatternFill("solid", fgColor=ACCENT)
            cell.font      = Font(bold=True, color=ACCENT_FG, size=10, name="Calibri")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        # Photo_Path — light yellow hint cell
        elif col_idx == 8:
            cell.fill      = PatternFill("solid", fgColor="FFFDE7")
            cell.font      = Font(size=10, name="Calibri", italic=True, color="9E9E9E")
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            if not value.strip():
                cell.value = "← paste photo path here"
        # Variation Count
        elif col_idx == 7:
            cell.fill      = PatternFill("solid", fgColor=bg)
            cell.font      = Font(size=10, name="Calibri", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            try:
                cell.value = int(value.strip())
            except ValueError:
                pass
        else:
            cell.fill      = PatternFill("solid", fgColor=bg)
            cell.font      = Font(size=10, name="Calibri")
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    ws.row_dimensions[row_idx].height = 72

# --- Column widths ---
for i, w in enumerate(col_widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Freeze header row
ws.freeze_panes = "A2"

# Auto-filter
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

wb.save(DST)
print(f"Saved: {DST}")
