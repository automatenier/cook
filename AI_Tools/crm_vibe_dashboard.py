import gspread
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# --- STYLING (VIBE MODE) ---
BLACK = "1A1A2E"
CARD_BG = "0F3460"
ACCENT = "E94560"
WHITE = "FFFFFF"
GRAY = "B0B0B0"
GREEN = "00C853"

def create_vibe_dashboard(data):
    wb = Workbook()
    ws = wb.active
    ws.title = "Vibe Dashboard"
    
    # 1. SETUP
    ws.sheet_properties.tabColor = ACCENT
    for row in range(1, 100):
        for col in range(1, 20):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill("solid", fgColor=BLACK)
            cell.font = Font(name="Aptos", color=WHITE)
            
    # 2. CALC METRICS
    total_leads = len(data)
    qualified = len([d for d in data if d.get('5-Status') == 'Qualified'])
    booked = len([d for d in data if str(d.get('9--Booked')).upper() == 'TRUE'])
    closed = len([d for d in data if d.get('11-Close') == 'Closed'])
    
    # 3. HEADER
    ws.merge_cells("B2:H2")
    title = ws["B2"]
    title.value = "CRM VIBE DASHBOARD (SAVASA)"
    title.font = Font(name="Aptos", size=24, bold=True, color=ACCENT)
    title.alignment = Alignment(horizontal="center")
    
    # 4. KPI CARDS
    def make_card(start_col, label, value):
        ws.merge_cells(start_row=4, start_column=start_col, end_row=4, end_column=start_col+1)
        ws.merge_cells(start_row=5, start_column=start_col, end_row=6, end_column=start_col+1)
        
        l_cell = ws.cell(row=4, column=start_col, value=label)
        l_cell.font = Font(name="Aptos", size=10, color=GRAY)
        l_cell.fill = PatternFill("solid", fgColor=CARD_BG)
        l_cell.alignment = Alignment(horizontal="center")
        
        v_cell = ws.cell(row=5, column=start_col, value=value)
        v_cell.font = Font(name="Aptos", size=32, bold=True, color=WHITE)
        v_cell.fill = PatternFill("solid", fgColor=CARD_BG)
        v_cell.alignment = Alignment(horizontal="center", vertical="center")
        
    make_card(2, "TOTAL LEADS", total_leads)
    make_card(4, "QUALIFIED", qualified)
    make_card(6, "BOOKED", booked)
    make_card(8, "CLOSED", closed)
    
    # 5. DATA TABLE
    ws.cell(row=8, column=2, value="RECENT ACTIVITY").font = Font(bold=True, size=12, color=ACCENT)
    headers = ["Name", "Status", "Booked", "Close", "Lead Date"]
    for i, h in enumerate(headers):
        cell = ws.cell(row=9, column=2+i, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor=CARD_BG)
        
    for idx, d in enumerate(data[-10:]): # Last 10
        ws.cell(row=10+idx, column=2, value=d.get('3-Username'))
        ws.cell(row=10+idx, column=3, value=d.get('5-Status'))
        ws.cell(row=10+idx, column=4, value=d.get('9--Booked'))
        ws.cell(row=10+idx, column=5, value=d.get('11-Close'))
        ws.cell(row=10+idx, column=6, value=d.get('13-Leadate'))
        
    filename = "PDCT_JO_Consult/F. Data Reporting/CRM_Vibe_Report.xlsx"
    wb.save(filename)
    return filename

if __name__ == "__main__":
    SERVICE_ACCOUNT_FILE = "service_account.json"
    SID = "1vJn-5RMcIxZLZBFoSlQJpOwcA_BxXdb39XXY_X2Mos8"
    gc = gspread.service_account(filename=SERVICE_ACCOUNT_FILE)
    sh = gc.open_by_key(SID)
    ws = sh.get_worksheet(0)
    data = ws.get_all_records()
    
    file = create_vibe_dashboard(data)
    print(f"Dashboard generated: {file}")
