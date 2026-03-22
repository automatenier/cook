#!/usr/bin/env python3
"""
CRM Manager — Unified interface for JO Consult CRM & Tracking infrastructure.

Handles creation of local Excel trackers and cloud-based Google Sheets systems.
This tool consolidates create_crm_system.py, create_excel_crm.py, and create_google_sheets.py.

Commands:
    py -3 AI_Tools/crm_manager.py --type excel [--output path/to/file.xlsx]
    py -3 AI_Tools/crm_manager.py --type google-3-system
    py -3 AI_Tools/crm_manager.py --type google-single
"""

import os
import sys
import json
import argparse
from pathlib import Path
from dotenv import load_dotenv, set_key

# Load environment variables
ENV_PATH = Path(__file__).parent.parent / ".env"
load_dotenv(ENV_PATH)

# --- Color Palette ---
C = {
    "green":        {"red": 0.69,  "green": 0.95,  "blue": 0.73},        
    "light_green":  {"red": 0.85,  "green": 0.98,  "blue": 0.87},        
    "orange":       {"red": 1.0,   "green": 0.85,  "blue": 0.66},        
    "light_orange": {"red": 1.0,   "green": 0.93,  "blue": 0.82},        
    "blue":         {"red": 0.64,  "green": 0.84,  "blue": 1.0},
    "light_blue":   {"red": 0.82,  "green": 0.92,  "blue": 1.0},
    "purple":       {"red": 0.82,  "green": 0.75,  "blue": 1.0},
    "light_purple": {"red": 0.91,  "green": 0.87,  "blue": 1.0},
    "red":          {"red": 1.0,   "green": 0.79,  "blue": 0.79},        
    "yellow":       {"red": 1.0,   "green": 0.96,  "blue": 0.70},        
    "dark_blue":    {"red": 0.24,  "green": 0.52,  "blue": 0.78},        
    "white":        {"red": 1.0,   "green": 1.0,   "blue": 1.0},
    "light_gray":   {"red": 0.95,  "green": 0.95,  "blue": 0.95},        
}

# --- Google Sheets Systems Definitions (From create_crm_system.py) ---
G_SYSTEMS_3 = [
    {
        "title": "JO Consult — CRM & Outreach",
        "env_key": "GOOGLE_SHEETS_CRM_ID",
        "sheets": [
            {
                "name": "Leads",
                "tab_color": C["green"],
                "header_color": C["green"],
                "headers": ["lead_id", "date_added", "lead_age_days", "username", "full_name", "wa_number", "email", "lead_type", "lead_source", "scrape_method", "setter", "status", "24_convo", "24_convo_date", "booked", "call_date", "showed", "no_show", "close_status", "close_date", "lead_to_close_days", "package", "cash_collected", "ltv", "commission", "mindmap", "objections", "objection_type", "content_feedback", "proposed", "live_calls_date", "notes"],
                "seed_row": {"C": '=IF(B2="","",TODAY()-B2)'}
            },
            {
                "name": "Outreach_DB",
                "tab_color": C["light_green"],
                "header_color": C["light_green"],
                "headers": ["profile_id", "scrape_date", "scrape_method", "username", "full_name", "bio", "followers", "following", "posts_count", "avg_views", "engagement_rate", "profile_url", "email", "wa_number", "external_url", "niche", "content_type", "is_verified", "is_business", "icp_score", "status", "added_to_crm", "crm_username", "notes"]
            },
            {
                "name": "Script_Performance",
                "tab_color": C["yellow"],
                "header_color": C["yellow"],
                "headers": ["script_id", "script_name", "version", "target_type", "script_preview", "date_created", "date_last_used", "sent_count", "reply_count", "reply_rate", "booked_count", "book_rate", "close_count", "close_rate", "top_objection", "status", "notes"],
                "seed_row": {"J": '=IF(H2=0,"",I2/H2)', "L": '=IF(I2=0,"",K2/I2)', "N": '=IF(K2=0,"",M2/K2)'}
            },
            {
                "name": "Dashboard_CRM",
                "tab_color": C["dark_blue"],
                "header_color": C["blue"],
                "dashboard": True,
                "layout": [
                    ["JO CONSULT — CRM DASHBOARD"],
                    [],
                    ["SETTER METRICS", "", "Total Leads", "", "Booked", "", "Showed"],
                    ["", "", "=COUNTA(Leads!D2:D1000)", "", "=COUNTIFS(Leads!O2:O1000,TRUE)", "", "=COUNTIFS(Leads!Q2:Q1000,TRUE)"],
                    [],
                    ["Sets Closed", "", "Close Rate", "", "Show Up Rate", "", "Total Revenue"],
                    ["=COUNTIFS(Leads!S2:S1000,\"Closed\")", "", "=IFERROR(COUNTIFS(Leads!S2:S1000,\"Closed\")/COUNTIFS(Leads!Q2:Q1000,TRUE),0)", "", "=IFERROR(COUNTIFS(Leads!Q2:Q1000,TRUE)/COUNTIFS(Leads!O2:O1000,TRUE),0)", "", "=SUM(Leads!W2:W1000)"]
                ]
            }
        ]
    },
    {
        "title": "JO Consult — Team EOD",
        "env_key": "GOOGLE_SHEETS_EOD_ID",
        "sheets": [
            {
                "name": "Founder_EOD",
                "tab_color": C["purple"],
                "header_color": C["purple"],
                "headers": ["date", "hours_worked", "scripts_written", "captions_written", "hooks_written", "agents_managed", "client_check_ins", "issues_resolved", "strategy_sessions", "deliverables_completed", "what_went_well", "improve", "energy_score", "focus_score", "notes"]
            },
            {
                "name": "Setter_EOD",
                "tab_color": C["green"],
                "header_color": C["green"],
                "headers": ["date", "setter_name", "productivity_score", "hours_worked", "inbox_checked", "new_convos", "follow_ups_sent", "yt_longform_sent", "calls_proposed", "calls_booked_dm", "qualified_bookings", "unqualified_bookings", "icp_leads_today", "convos_nurtured", "calls_on_calendar", "calls_showed", "sets_closed", "revenue_closed", "what_went_well", "improve", "situation_update", "top_objection", "content_feedback", "notes"]
            }
        ]
    }
]

# --- Helpers ---
def col_to_letter(n: int) -> str:
    result = ""
    while n > 0:
        n -= 1
        result = chr(n % 26 + ord("A")) + result
        n //= 26
    return result

def letter_to_col(s: str) -> int:
    result = 0
    for ch in s.upper():
        result = result * 26 + (ord(ch) - ord("A") + 1)
    return result

def setup_oauth():
    client_id = os.getenv("GOOGLE_CLIENT_ID")
    client_secret = os.getenv("GOOGLE_CLIENT_SECRET")
    if not client_id or not client_secret:
        print("ERROR: GOOGLE_CLIENT_ID / GOOGLE_CLIENT_SECRET missing from .env")
        sys.exit(1)

    creds_dir = Path.home() / ".config" / "gspread"
    creds_dir.mkdir(parents=True, exist_ok=True)
    with open(creds_dir / "credentials.json", "w") as f:
        json.dump({
            "installed": {
                "client_id": client_id,
                "client_secret": client_secret,
                "redirect_uris": ["http://localhost"],
                "auth_uri": "https://accounts.google.com/o/oauth2/auth", 
                "token_uri": "https://oauth2.googleapis.com/token",      
            }
        }, f)

def build_google_sheet(ws, sheet_def):
    is_dashboard = sheet_def.get("dashboard", False)
    header_color = sheet_def.get("header_color", {"red": 0.9, "green": 0.9, "blue": 0.9})

    if is_dashboard:
        layout = sheet_def.get("layout", [])
        if not layout: return
        max_cols = max((len(r) for r in layout), default=1)
        data = [list(row) + [""] * (max_cols - len(row)) for row in layout]
        ws.update("A1", data, value_input_option="USER_ENTERED")
        ws.format(f"A1:{col_to_letter(max_cols)}1", {"textFormat": {"bold": True, "fontSize": 13}, "backgroundColor": header_color})
        ws.freeze(rows=1)
    else:
        headers = sheet_def.get("headers", [])
        if not headers: return
        ws.update("A1", [headers], value_input_option="USER_ENTERED")
        ws.format(f"A1:{col_to_letter(len(headers))}1", {"textFormat": {"bold": True, "fontSize": 10}, "backgroundColor": header_color})
        ws.freeze(rows=1)
        seed = sheet_def.get("seed_row", {})
        for col_letter, formula in seed.items():
            ws.update_cell(2, letter_to_col(col_letter), formula)

# --- Execution ---
def handle_excel(output_path):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("Error: openpyxl not installed.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"
    headers = ["ID", "Date Added", "Username", "Name", "Platform", "Lead Type", "Source", "Stage", "Temperature", "Setter", "Notes"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    if not output_path:
        output_path = Path(__file__).parent.parent / ".tmp" / "JO_Consult_CRM.xlsx"
    else:
        output_path = Path(output_path)
    
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Excel CRM created at: {output_path}")

def handle_google(system_type):
    try:
        import gspread
    except ImportError:
        print("Error: gspread not installed.")
        return

    setup_oauth()
    gc = gspread.oauth()
    
    if system_type == "google-single":
        sh = gc.create("JO Consult - Data System")
        print(f"Created Single Google Sheet: {sh.url}")
        set_key(str(ENV_PATH), "GOOGLE_SHEETS_CRM_ID", sh.id)
    else:
        for sys_def in G_SYSTEMS_3:
            print(f"Creating: {sys_def['title']}")
            sh = gc.create(sys_def["title"])
            first = True
            for sheet_def in sys_def["sheets"]:
                sname = sheet_def["name"]
                if first:
                    ws = sh.sheet1
                    ws.update_title(sname)
                    first = False
                else:
                    ws = sh.add_worksheet(title=sname, rows=1000, cols=30)
                build_google_sheet(ws, sheet_def)
                print(f"  ✓ {sname}")
            set_key(str(ENV_PATH), sys_def["env_key"], sh.id)
            print(f"  → {sh.url}\n")

def main():
    parser = argparse.ArgumentParser(description="JO Consult CRM Manager")
    parser.add_argument("--type", choices=["excel", "google-3-system", "google-single"], required=True, help="Type of CRM to create")
    parser.add_argument("--output", help="Output path for Excel file")
    args = parser.parse_args()
    
    if args.type == "excel":
        handle_excel(args.output)
    else:
        handle_google(args.type)

if __name__ == "__main__":
    main()
