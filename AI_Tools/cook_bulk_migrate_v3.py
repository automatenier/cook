#!/usr/bin/env python3
"""
Bulk Migrate Excel files using Service Account into a SPECIFIC Folder.
Bypasses "Quota Exceeded" by using the parent folder's ownership/storage.
"""
import json
import os
import sys
import argparse
from pathlib import Path
import openpyxl
import gspread
from datetime import date, datetime

# Paths
ROOT = Path(__file__).parent.parent
SERVICE_ACCOUNT_FILE = ROOT / "service_account.json"
MEMORY_PATH = ROOT / "AI_MEMORY" / "google_sheets_map.json"

def to_sheets_value(v):
    if v is None: return ""
    if isinstance(v, str):
        stripped = v.strip()
        if stripped.startswith("="): return stripped
        return stripped
    if isinstance(v, (datetime, date)):
        return v.strftime("%d/%m/%Y")
    if isinstance(v, bool): return v
    return v

def last_data_row(ws) -> int:
    for r in range(ws.max_row, 0, -1):
        if any(ws.cell(r, c).value is not None for c in range(1, ws.max_column + 1)):
            return r
    return 1

def migrate_file(gc, file_path, folder_id):
    print(f"Migrating: {file_path.name}...")
    try:
        wb = openpyxl.load_workbook(file_path, data_only=False)
    except Exception as e:
        print(f"  FAILED to read: {e}")
        return None

    title = f"Cook — {file_path.stem}"
    try:
        # Create the spreadsheet in the specific folder
        # Note: gspread's create() doesn't support folder_id directly in all versions, 
        # but we can move it or use the drive API via gc.auth
        sh = gc.create(title, folder_id=folder_id)
        print(f"  Created in folder: {folder_id}")
    except Exception as e:
        print(f"  FAILED to create Google Sheet: {e}")
        # Fallback for older gspread: create then move
        try:
            sh = gc.create(title)
            # Move logic would go here, but usually quota fails before this.
            # If the error was 403 quota, this fallback won't help.
        except:
            return None

    first = True
    for sheet_name in wb.sheetnames:
        ws_xl = wb[sheet_name]
        max_col = ws_xl.max_column
        end_row = last_data_row(ws_xl)

        if first:
            ws_gs = sh.sheet1
            ws_gs.update_title(sheet_name)
            first = False
        else:
            ws_gs = sh.add_worksheet(title=sheet_name, rows=max(end_row+20, 100), cols=max(max_col+2, 20))

        data = []
        for r in range(1, end_row + 1):
            row = [to_sheets_value(ws_xl.cell(r, c).value) for c in range(1, max_col + 1)]
            data.append(row)

        if data:
            ws_gs.update("A1", data, value_input_option="USER_ENTERED")
        
        ws_gs.format("A1:Z1", {"textFormat": {"bold": True}})
        ws_gs.freeze(rows=1)
        print(f"  ✓ {sheet_name}")

    return {"id": sh.id, "url": f"https://docs.google.com/spreadsheets/d/{sh.id}"}

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--dir", required=True)
    parser.add_argument("--folder", required=True, help="Google Drive Folder ID")
    parser.add_argument("--share", help="Email to share all sheets with")
    args = parser.parse_args()

    if not SERVICE_ACCOUNT_FILE.exists():
        print(f"Error: {SERVICE_ACCOUNT_FILE} not found.")
        return

    gc = gspread.service_account(filename=str(SERVICE_ACCOUNT_FILE))
    
    source_dir = Path(args.dir)
    xlsx_files = list(source_dir.glob("*.xlsx"))
    
    if MEMORY_PATH.exists():
        with open(MEMORY_PATH, "r") as f:
            sheet_map = json.load(f)
    else:
        sheet_map = {}

    for xlsx in xlsx_files:
        if xlsx.name in sheet_map:
            print(f"Skipping {xlsx.name}")
            continue
            
        result = migrate_file(gc, xlsx, args.folder)
        if result:
            sheet_map[xlsx.name] = result
            if args.share:
                try:
                    sh = gc.open_by_key(result["id"])
                    sh.share(args.share, perm_type='user', role='writer')
                    print(f"  Shared with {args.share}")
                except Exception as e:
                    print(f"  Failed to share: {e}")

            MEMORY_PATH.parent.mkdir(parents=True, exist_ok=True)
            with open(MEMORY_PATH, "w") as f:
                json.dump(sheet_map, f, indent=2)

    print(f"\nMigration complete. Map saved to {MEMORY_PATH}")

if __name__ == "__main__":
    main()
