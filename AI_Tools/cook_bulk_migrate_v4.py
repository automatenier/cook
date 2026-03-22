#!/usr/bin/env python3
"""
Bulk Migrate Excel files using Service Account into a SPECIFIC Folder.
Using Drive API directly to ensure ownership/quota is handled correctly.
"""
import json
import os
import sys
import argparse
from pathlib import Path
import openpyxl
import gspread
from googleapiclient.discovery import build
from google.oauth2 import service_account
from datetime import date, datetime

# Paths
ROOT = Path(__file__).parent.parent
SERVICE_ACCOUNT_FILE = ROOT / "service_account.json"
MEMORY_PATH = ROOT / "AI_MEMORY" / "google_sheets_map.json"

SCOPES = [
    'https://www.googleapis.com/auth/spreadsheets',
    'https://www.googleapis.com/auth/drive'
]

def to_sheets_value(v):
    if v is None: return ""
    if isinstance(v, str):
        stripped = v.strip()
        if stripped.startswith("="): return stripped
        return stripped
    if isinstance(v, (datetime, date)):
        return v.strftime("%d/%m/%Y")
    if isinstance(v, bool): return v
    return v

def last_data_row(ws) -> int:
    for r in range(ws.max_row, 0, -1):
        if any(ws.cell(r, c).value is not None for c in range(1, ws.max_column + 1)):
            return r
    return 1

def migrate_file(gc, drive_service, file_path, folder_id):
    print(f"Migrating: {file_path.name}...")
    try:
        wb = openpyxl.load_workbook(file_path, data_only=False)
    except Exception as e:
        print(f"  FAILED to read Excel: {e}")
        return None

    title = f"Cook — {file_path.stem}"
    
    # Step 1: Create the empty spreadsheet in the folder using Drive API
    file_metadata = {
        'name': title,
        'parents': [folder_id],
        'mimeType': 'application/vnd.google-apps.spreadsheet'
    }
    try:
        res = drive_service.files().create(body=file_metadata, fields='id').execute()
        spreadsheet_id = res.get('id')
        print(f"  Created Spreadsheet ID: {spreadsheet_id}")
    except Exception as e:
        print(f"  FAILED to create in Drive: {e}")
        return None

    # Step 2: Open with gspread to fill data
    sh = gc.open_by_key(spreadsheet_id)

    first = True
    for sheet_name in wb.sheetnames:
        ws_xl = wb[sheet_name]
        max_col = ws_xl.max_column
        end_row = last_data_row(ws_xl)

        if first:
            ws_gs = sh.sheet1
            ws_gs.update_title(sheet_name)
            first = False
        else:
            ws_gs = sh.add_worksheet(title=sheet_name, rows=max(end_row+20, 100), cols=max(max_col+2, 20))

        data = []
        for r in range(1, end_row + 1):
            row = [to_sheets_value(ws_xl.cell(r, c).value) for c in range(1, max_col + 1)]
            data.append(row)

        if data:
            ws_gs.update("A1", data, value_input_option="USER_ENTERED")
        
        ws_gs.format("A1:Z1", {"textFormat": {"bold": True}})
        ws_gs.freeze(rows=1)
        print(f"  ✓ {sheet_name}")

    return {"id": sh.id, "url": f"https://docs.google.com/spreadsheets/d/{sh.id}"}

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--dir", required=True)
    parser.add_argument("--folder", required=True)
    parser.add_argument("--share", help="Email to share with")
    args = parser.parse_args()

    if not SERVICE_ACCOUNT_FILE.exists():
        print(f"Error: {SERVICE_ACCOUNT_FILE} not found.")
        return

    # Auth for gspread
    gc = gspread.service_account(filename=str(SERVICE_ACCOUNT_FILE))
    
    # Auth for Google Drive API
    creds = service_account.Credentials.from_service_account_file(
        str(SERVICE_ACCOUNT_FILE), scopes=SCOPES)
    drive_service = build('drive', 'v3', credentials=creds)

    source_dir = Path(args.dir)
    xlsx_files = list(source_dir.glob("*.xlsx"))
    
    if MEMORY_PATH.exists():
        with open(MEMORY_PATH, "r") as f:
            sheet_map = json.load(f)
    else:
        sheet_map = {}

    for xlsx in xlsx_files:
        if xlsx.name in sheet_map:
            print(f"Skipping {xlsx.name}")
            continue
            
        result = migrate_file(gc, drive_service, xlsx, args.folder)
        if result:
            sheet_map[xlsx.name] = result
            if args.share:
                try:
                    # Drive API sharing is more robust for service accounts
                    batch_permission = {
                        'type': 'user',
                        'role': 'writer',
                        'emailAddress': args.share
                    }
                    drive_service.permissions().create(
                        fileId=result["id"],
                        body=batch_permission,
                        fields='id',
                    ).execute()
                    print(f"  Shared with {args.share}")
                except Exception as e:
                    print(f"  Failed to share: {e}")

            MEMORY_PATH.parent.mkdir(parents=True, exist_ok=True)
            with open(MEMORY_PATH, "w") as f:
                json.dump(sheet_map, f, indent=2)

    print(f"\nMigration complete. Map saved to {MEMORY_PATH}")

if __name__ == "__main__":
    main()
