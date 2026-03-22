import openpyxl
import sys
import os

def peek_excel(file_path, rows=5):
    if not os.path.exists(file_path):
        print(f"Error: File '{file_path}' not found.")
        return

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        for sheet_name in wb.sheetnames:
            print(f"\n--- Sheet: {sheet_name} ---")
            sheet = wb[sheet_name]
            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                if i >= rows:
                    break
                print(row)
    except Exception as e:
        print(f"Error processing {file_path}: {str(e)}")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python peek_excel.py <file_path>")
    else:
        peek_excel(sys.argv[1])
