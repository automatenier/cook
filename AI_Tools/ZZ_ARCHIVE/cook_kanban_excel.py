"""
cook_kanban_excel.py — Generate visual Kanban board in Excel

Output: 01 HMN__Command/00_COOK/KANBAN.xlsx
Usage:  py -3 AI_Tools/cook_kanban_excel.py

Prefix color legend:
  [WRITE]   light blue    — scripting, copy, course content
  [RENDER]  light purple  — Remotion, ElevenLabs, video export
  [REVIEW]  light orange  — anything needing eyes/approval
  [ADMIN]   light green   — invoices, agreements, onboarding
  [TOOL]    light gray    — Python scripts, n8n automations
  [CALL]    light yellow  — calls, interviews, discovery
  [PLAN]    light teal    — strategy, research, roadmap
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT        = Path(__file__).resolve().parent.parent
OUTPUT_PATH = ROOT / "01 HMN__Command" / "00_COOK" / "KANBAN.xlsx"

# ── Prefix styling ───────────────────────────────────────────────────────────
PREFIX_STYLE = {
    "[WRITE]":  {"bg": "D6E4F7", "fg": "1F4E79"},
    "[RENDER]": {"bg": "EAD1FA", "fg": "5B0F8A"},
    "[REVIEW]": {"bg": "FDE8C8", "fg": "7D3C0A"},
    "[ADMIN]":  {"bg": "D5F0D5", "fg": "1E5631"},
    "[TOOL]":   {"bg": "E8E8E8", "fg": "404040"},
    "[CALL]":   {"bg": "FFF3CC", "fg": "7D6608"},
    "[PLAN]":   {"bg": "D5EEF0", "fg": "0E5A6B"},
}

# ── Lane data ────────────────────────────────────────────────────────────────
LANES = [
    {
        "name":  "📥  INBOX",
        "color": "2E4057",
        "tasks": [
            "[TOOL] Run /inbox-triage — check new inputs",
            "[ADMIN] Fadli — Confirm welcome email sent",
            "[ADMIN] Mathew Jordan — Confirm niche, sprint start, payment",
            "[ADMIN] Ruth — Confirm niche, sprint start, payment",
        ],
    },
    {
        "name":  "🗓️  THIS WEEK",
        "color": "C55A11",
        "tasks": [
            "[RENDER] Fadli — Batch 1-10 March 2026 Sprint",
            "[RENDER] Mathew Jordan — Cikarang Tokyo + March Sprint",
            "[RENDER] Ruth — March Sprint script",
            "[RENDER] Real Estate — March Sprint script",
            "[CALL] Fadli — Book onboarding call",
            "[ADMIN] Fadli — Collect 10 min voice recording",
            "[ADMIN] Fadli — Invoice + Legal Agreement + Guarantee",
            "[TOOL] Fadli — Run offer_agent.py → Offer Sheet",
            "[WRITE] Fadli — VSL script draft",
            "[WRITE] Fadli — Reactivation scripts",
            "[CALL] Brighton — STAR answers (ROAS 413x + 6 units)",
        ],
    },
    {
        "name":  "⚙️  IN PROGRESS",
        "color": "7030A0",
        "tasks": [
            "[RENDER] Fadli — 5x-rule-fitness (2026-03-02)",
            "[PLAN] SAVASA — Interview prep (USP, Talking Points)",
        ],
    },
    {
        "name":  "👀  REVIEW / WAITING",
        "color": "BF8F00",
        "tasks": [
            "[ADMIN] Fadli — Payment confirmation (waiting n8n/Jordan)",
        ],
    },
    {
        "name":  "✅  DONE",
        "color": "375623",
        "tasks": [
            "[ADMIN] Fadli — Onboarding email sent (akses + komunitas)",
        ],
    },
]

LEGEND_ROWS = [
    ("[WRITE]",  "Script writing, copy, course content, offer sheets"),
    ("[RENDER]", "Remotion render, ElevenLabs voiceover, video export"),
    ("[REVIEW]", "Anything waiting on your eyes or approval"),
    ("[ADMIN]",  "Invoices, agreements, onboarding setup, confirmations"),
    ("[TOOL]",   "Running a Python script or n8n automation"),
    ("[CALL]",   "Calls, interviews, discovery sessions"),
    ("[PLAN]",   "Strategy, research, roadmap, curriculum structure"),
]

# ── Helpers ──────────────────────────────────────────────────────────────────
def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _border(color: str = "D0D0D0") -> Border:
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _font(color="000000", size=10, bold=False, italic=False) -> Font:
    return Font(color=color, size=size, bold=bold, italic=italic, name="Segoe UI")

def _get_prefix(task: str) -> str:
    for p in PREFIX_STYLE:
        if task.startswith(p):
            return p
    return ""


# ── Build ────────────────────────────────────────────────────────────────────
def build_kanban_sheet(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Kanban"
    ws.sheet_view.showGridLines = False

    LANE_WIDTH = 42
    GAP_WIDTH  = 2

    # Column layout: lanes at cols 1, 3, 5, 7, 9 — gaps between
    lane_cols = []
    for i in range(len(LANES)):
        col = i * 2 + 1
        lane_cols.append(col)
        ws.column_dimensions[get_column_letter(col)].width = LANE_WIDTH
        if i < len(LANES) - 1:
            ws.column_dimensions[get_column_letter(col + 1)].width = GAP_WIDTH

    # ── Row 1: Lane headers ──────────────────────────────────────────────────
    ws.row_dimensions[1].height = 32
    for lane, col in zip(LANES, lane_cols):
        c = ws.cell(row=1, column=col, value=lane["name"])
        c.fill      = _fill(lane["color"])
        c.font      = _font(color="FFFFFF", size=12, bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = _border("FFFFFF")

    # ── Row 2: Task count badge ──────────────────────────────────────────────
    ws.row_dimensions[2].height = 16
    for lane, col in zip(LANES, lane_cols):
        n = len(lane["tasks"])
        c = ws.cell(row=2, column=col, value=f"{n} task{'s' if n != 1 else ''}")
        c.fill      = _fill("F0F0F0")
        c.font      = _font(color="999999", size=9, italic=True)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = _border()

    # ── Rows 3+: Cards ───────────────────────────────────────────────────────
    max_tasks = max(len(lane["tasks"]) for lane in LANES)

    for row_offset in range(max_tasks):
        row = row_offset + 3
        ws.row_dimensions[row].height = 38

        for lane, col in zip(LANES, lane_cols):
            if row_offset < len(lane["tasks"]):
                task   = lane["tasks"][row_offset]
                prefix = _get_prefix(task)
                style  = PREFIX_STYLE.get(prefix, {"bg": "F5F5F5", "fg": "333333"})

                c = ws.cell(row=row, column=col, value=task)
                c.fill      = _fill(style["bg"])
                c.font      = _font(color=style["fg"], size=10)
                c.alignment = Alignment(wrap_text=True, vertical="top", indent=1)
                c.border    = _border()
            else:
                c = ws.cell(row=row, column=col)
                c.fill   = _fill("FAFAFA")
                c.border = _border("EEEEEE")

    ws.freeze_panes = "A3"


def build_legend_sheet(wb: Workbook) -> None:
    ls = wb.create_sheet("Legend")
    ls.sheet_view.showGridLines = False
    ls.column_dimensions["A"].width = 14
    ls.column_dimensions["B"].width = 52

    # Header
    ls.row_dimensions[1].height = 26
    for col, label in [(1, "PREFIX"), (2, "WORK TYPE")]:
        c = ls.cell(row=1, column=col, value=label)
        c.fill      = _fill("2E4057")
        c.font      = _font(color="FFFFFF", size=11, bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = _border("FFFFFF")

    for i, (prefix, desc) in enumerate(LEGEND_ROWS, start=2):
        ls.row_dimensions[i].height = 22
        style = PREFIX_STYLE.get(prefix, {"bg": "F5F5F5", "fg": "333333"})

        ca = ls.cell(row=i, column=1, value=prefix)
        ca.fill      = _fill(style["bg"])
        ca.font      = _font(color=style["fg"], size=10, bold=True)
        ca.alignment = Alignment(horizontal="center", vertical="center")
        ca.border    = _border()

        cb = ls.cell(row=i, column=2, value=desc)
        cb.font      = _font(color="333333", size=10)
        cb.alignment = Alignment(vertical="center", indent=1)
        cb.border    = _border()


def build() -> None:
    wb = Workbook()
    build_kanban_sheet(wb)
    build_legend_sheet(wb)

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"[OK] Kanban saved -> {OUTPUT_PATH}")


if __name__ == "__main__":
    build()
