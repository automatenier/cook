import openpyxl
import os
import sys

def check_excel_health(file_path):
    errors = []
    warnings = []
    
    if not os.path.exists(file_path):
        return [f"ERROR: File not found: {file_path}"], []

    try:
        # Load workbook with data_only=True to see the values (including formula errors)
        wb = openpyxl.load_workbook(file_path, data_only=True)
        
        if not wb.sheetnames:
            errors.append("ERROR: Workbook has no sheets.")
            return errors, warnings

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            
            # Check if sheet is empty
            if sheet.max_row <= 1 and sheet.max_column <= 1 and sheet.cell(row=1, column=1).value is None:
                warnings.append(f"Sheet '{sheet_name}' appears to be empty.")
                continue

            # Check for formula errors in all cells
            for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
                for col_idx, value in enumerate(row, 1):
                    if isinstance(value, str) and value.startswith("#") and any(err in value for err in ["#REF!", "#VALUE!", "#N/A", "#NAME?", "#DIV/0!", "#NULL!", "#NUM!"]):
                        errors.append(f"[{sheet_name}] Formula Error '{value}' at {openpyxl.utils.get_column_letter(col_idx)}{row_idx}")

            # Logical Checks for CRM-like sheets (Leads, CRM, Tracker)
            if any(kw in sheet_name.lower() for kw in ["lead", "crm", "tracker", "onboard"]):
                headers = [str(cell.value).lower() if cell.value else "" for cell in sheet[1]]
                
                # Check for critical missing headers
                critical_headers = ["name", "status", "date"]
                found_headers = [h for h in headers if any(ch in h for ch in critical_headers)]
                if not found_headers and sheet.max_row > 1:
                    warnings.append(f"[{sheet_name}] Potential missing critical headers (Name, Status, Date).")

                # Logic: If 'Booked' is checked but no 'Date'
                col_map = {h: i for i, h in enumerate(headers)}
                booked_col = next((i for i, h in enumerate(headers) if "booked" in h), None)
                date_col = next((i for i, h in enumerate(headers) if "date" in h), None)
                name_col = next((i for i, h in enumerate(headers) if "name" in h or "username" in h), None)

                if booked_col is not None and date_col is not None:
                    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
                        booked_val = row[booked_col]
                        date_val = row[date_col]
                        name_val = row[name_col] if name_col is not None else f"Row {row_idx}"
                        
                        if booked_val in [True, "TRUE", "Yes", "yes", 1] and not date_val:
                            warnings.append(f"[{sheet_name}] Logical Issue: '{name_val}' is marked as Booked but has no Date at row {row_idx}.")

        return errors, warnings

    except Exception as e:
        return [f"ERROR: Failed to process {file_path}: {str(e)}"], []

def main():
    files = [
        "01 HMN__Command/00_AGENT_COMMAND.xlsx",
        "01 HMN__Command/00_COMMAND_CENTER.xlsx",
        "01 HMN__Command/00_CONTENT/Authority_Proof_Planner.xlsx",
        "01 HMN__Command/00_CONTENT/Content_Planner.xlsx",
        "01 HMN__Command/00_CONTENT/Content_Tracker.xlsx",
        "01 HMN__Command/00_CONTENT/reels_analysis_31.xlsx",
        "01 HMN__Command/00_COOK/Daily_KPI_Tracker_AgentKPI.xlsx",
        "01 HMN__Command/00_COOK/Daily_KPI_Tracker.xlsx",
        "01 HMN__Command/10_JO_Consult/Dash/JO_Consult_CRM.xlsx",
        "01 HMN__Command/10_JO_Consult/Dash/JO_Dashboard.xlsx",
        "01 HMN__Command/10_JO_Consult/Dash/JO_lead_magnet_tracker.xlsx",
        "01 HMN__Command/JO_Onboard.xlsx"
    ]

    print("# Excel Health Scan Report\n")
    
    total_errors = 0
    total_warnings = 0

    for f in files:
        # Use forward slashes for better cross-platform compatibility even on Windows with os.path.join
        full_path = os.path.join(os.getcwd(), f.replace("/", os.sep))
        print(f"## Scanning: {f}")
        errors, warnings = check_excel_health(full_path)
        
        if not errors and not warnings:
            print("  - ✅ No issues found.")
        else:
            for err in errors:
                print(f"  - ❌ {err}")
                total_errors += 1
            for warn in warnings:
                print(f"  - ⚠️ {warn}")
                total_warnings += 1
        print("")

    print(f"**Summary:** Found {total_errors} errors and {total_warnings} warnings across {len(files)} files.")

if __name__ == "__main__":
    main()
