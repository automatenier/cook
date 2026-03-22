"""
Generate the 14-Day Client Sprint SOP as a formatted Excel file.

Outputs a single .xlsx with two sheets:
  - Sprint SOP   : full day-by-day task table
  - Legend       : tag definitions + column guide

Usage:
    py -3 AI_Tools/generate_sprint_sop.py
    py -3 AI_Tools/generate_sprint_sop.py --output .tmp/Sprint_SOP.xlsx
"""

import argparse
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: py -3 -m pip install openpyxl")
    sys.exit(1)

# ---------------------------------------------------------------------------
# DATA
# ---------------------------------------------------------------------------

TASKS = [
    # (day, phase_label, task_num, task_name, tag, who, description, output_file, depends_on)
    # DAY 0
    ("0", "Client Signed", 1, "Create Folder Structure",           "🤖 AUTO",    "Claude",  "Run bash to create all client subfolders on OneDrive",                            "OneDrive folder tree",                        "Client profile filled"),
    ("0", "Client Signed", 2, "Generate Welcome Email",             "🤖 AUTO",    "Claude",  "Draft full welcome email in Indonesian using template + client profile",           "_CLIENT_DELIVERABLES/[C]/welcome_email.md",   "Google Calendar link"),
    ("0", "Client Signed", 3, "Generate Asset Collection Message",  "🤖 AUTO",    "Claude",  "WhatsApp/Telegram message listing all assets client needs to send",               "Ready-to-copy WA message",                    "Client profile filled"),
    ("0", "Client Signed", 4, "Create Google Calendar Booking",     "📥 COLLECT", "Jordan",  "Create 60-min appointment schedule for Day 14 consultation",                      "Booking link pasted in worksheet",            "—"),
    ("0", "Client Signed", 5, "Create GDrive Shared Folder",        "📥 COLLECT", "Jordan",  "Create 'Client — Assets' folder with 5 subfolders, invite client",               "GDrive link pasted in worksheet",             "—"),
    ("0", "Client Signed", 6, "Community + PDCT_MODULES Access",          "⚙️ PLATFORM","Jordan",  "Add client to WhatsApp group, Telegram community, share module access link",      "Confirmed in worksheet",                      "—"),
    ("0", "Client Signed", 7, "Telegram Progress Channel (n8n)",    "⚙️ PLATFORM","Jordan",  "Trigger n8n to create dedicated Telegram progress channel for client",            "Channel link pasted in worksheet",            "n8n live"),

    # DAY 1
    ("1", "Onboarding Kickoff", 1, "Generate Onboarding Instructions", "🤖 AUTO",    "Claude",  "Message + guide for client: how to fill forms, upload voice, upload brand assets", "Ready-to-send message",                       "Day 0 complete"),
    ("1", "Onboarding Kickoff", 2, "Send Onboarding Form to Client",   "📥 COLLECT", "Jordan",  "Send onboarding_form_template.md to client; wait for completed answers",           "onboarding.json saved in _CLIENT_DELIVERABLES","—"),
    ("1", "Onboarding Kickoff", 3, "Collect Client Info Docs (x4)",    "📥 COLLECT", "Client",  "Client fills: life_story, product_market_fit, offer_positioning, produk.md",       "4 docs uploaded to GDrive",                   "GDrive folder live"),
    ("1", "Onboarding Kickoff", 4, "Collect Voice Sample",             "📥 COLLECT", "Client",  "5-10 min clean audio (no music/reverb) uploaded to GDrive Voice Sample subfolder", "Voice sample in GDrive",                      "GDrive folder live"),
    ("1", "Onboarding Kickoff", 5, "Collect Brand Assets",             "📥 COLLECT", "Client",  "Logo, brand colors, fonts, top 5 posts, testimonials uploaded to GDrive",          "Brand assets in GDrive",                      "GDrive folder live"),

    # DAY 2
    ("2", "Full Build", 1, "Process Client Info → Content Angles",  "🤖 AUTO",    "Claude",  "Extract 5 authenticity hooks, 5 value-CTA hooks, unique mechanism, avatar journey", "swipe_file.md + content_angles.md",           "ONBOARDING_JSON ✓ + CLIENT_INFO ✓"),
    ("2", "Full Build", 2, "Build Swipe File + Graphics Guide",     "🤖 AUTO",    "Claude",  "Full swipe file + color palette from onboarding + brand assets",                   "swipe_file.md + graphics/color_palette.md",   "ONBOARDING_JSON ✓"),
    ("2", "Full Build", 3, "Run Offer Agent → Generate Offer Sheet","🤖 AUTO",    "Claude",  "py -3 tools/offer_agent.py --from-file onboarding.json",                           "offer_sheet.md + offer_sheet.json",           "onboarding.json saved"),
    ("2", "Full Build", 4, "Generate Sales Call Script",            "🤖 AUTO",    "Claude",  "Read swipe file + offer sheet → full sales call script",                           "sales_call_script.md",                        "swipe_file.md + offer_sheet.md"),
    ("2", "Full Build", 5, "Generate 7-Email Nurture Sequence",     "🤖 AUTO",    "Claude",  "7 emails (Day 0,1,3,5,7,10,14) from swipe file + offer sheet",                    "email_sequence.md",                           "swipe_file.md + offer_sheet.md"),
    ("2", "Full Build", 6, "Generate Ad Copy (3 Platforms)",        "🤖 AUTO",    "Claude",  "FB (3) + IG (3) + LinkedIn (2) ad copy variants",                                 "ad_copy.md",                                  "swipe_file.md + offer_sheet.md"),
    ("2", "Full Build", 7, "Generate DFY Bonus Stack (5 Bonuses)",  "🤖 AUTO",    "Claude",  "Using dfy_deliverables_prompts.md PROMPT 1; full content for each bonus",          "bonus_1–5.md + bonus_stack_summary.md",       "swipe_file.md + offer_sheet.md + PMF"),
    ("2", "Full Build", 8, "Generate Transformation Roadmap",       "🤖 AUTO",    "Claude",  "4-phase roadmap using PROMPT 2 from dfy_deliverables_prompts.md",                  "roadmap.md",                                  "swipe_file.md + life_story + produk"),
    ("2", "Full Build", 9, "Jordan Reviews + Approves",             "📥 COLLECT", "Jordan",  "Review offer_sheet.md + swipe_file.md; approve or note revision requests",         "OFFER_SHEET_APPROVED ✓",                      "All 8 AUTO tasks done"),

    # DAY 3
    ("3", "Sales Toolkit + Voice", 1, "Export Offer Sheet → HTML",      "🤖 AUTO",    "Claude",  "Styled HTML for PDF print from offer_sheet.md",                                   "funnel/offer_sheet.html",                     "OFFER_SHEET_APPROVED ✓"),
    ("3", "Sales Toolkit + Voice", 2, "Generate VSL Script",            "🤖 AUTO",    "Claude",  "Hook→Story→Problem→Solution→Proof→Offer→CTA; teleprompter format",                "vsl_script.md",                               "swipe_file.md + offer_sheet.md + client_info"),
    ("3", "Sales Toolkit + Voice", 3, "Build Netlify VSL Funnel HTML",  "🤖 AUTO",    "Claude",  "Full 1-page funnel HTML with Wistia placeholder slots + brand colors",            "funnel/index.html (placeholder)",             "vsl_script.md + color_palette.md"),
    ("3", "Sales Toolkit + Voice", 4, "ElevenLabs Voice Clone Setup",   "⚙️ PLATFORM","Jordan",  "Upload voice sample → create Instant Voice Clone → copy Voice ID",               "ELEVENLABS_VOICE_ID pasted in worksheet",     "Voice sample in GDrive"),

    # DAY 4
    ("4", "Filming + Funnel Live", 1, "Generate Filming Shot List",       "🤖 AUTO",    "Claude",  "Talking head setups, b-roll, VSL recording, testimonials, asset photos",          "filming_brief.md",                            "swipe_file.md + offer_sheet.md"),
    ("4", "Filming + Funnel Live", 2, "Filming Session (2 hours)",        "🎬 ONSITE",  "Jordan",  "Talking head, b-roll, VSL recording, testimonials, asset photos → upload to GDrive","Raw footage in GDrive",                       "filming_brief.md printed"),
    ("4", "Filming + Funnel Live", 3, "Edit + Upload to Wistia",          "🎬 ONSITE",  "Jordan",  "Edit VSL + testimonial clips in CapCut → upload to Wistia → copy media IDs",      "WISTIA_IDs pasted in worksheet",              "Footage uploaded"),
    ("4", "Filming + Funnel Live", 4, "Finalize Netlify Funnel",          "🤖 AUTO",    "Claude",  "Replace placeholder Wistia IDs + insert calendar/WA links → final index.html",    "funnel/index.html (final)",                   "WISTIA_IDs collected"),
    ("4", "Filming + Funnel Live", 5, "Deploy to Netlify",                "⚙️ PLATFORM","Jordan",  "netlify deploy --prod --dir=funnel/",                                             "NETLIFY_URL pasted in worksheet",             "funnel/index.html final"),
    ("4", "Filming + Funnel Live", 6, "Booking Calendar + n8n Reminders", "⚙️ PLATFORM","Jordan",  "Connect n8n: new booking → WA reminder 24h + 1h before",                         "BOOKING_REMINDERS: live ✓",                   "Netlify live"),
    ("4", "Filming + Funnel Live", 7, "Test Checklist",                   "🤖 AUTO",    "Claude",  "Generate checklist for Jordan to verify live funnel",                             "Test checklist in chat",                      "Netlify live"),

    # DAY 5
    ("5", "Content Audit", 1, "Collect Last 20 Posts Data",        "📥 COLLECT", "Jordan",  "Pull last 20 IG posts: caption, views, reach, saves, comments, date",              "POSTS_DATA pasted or file in GDrive",         "—"),
    ("5", "Content Audit", 2, "Run Content Audit",                  "🤖 AUTO",    "Claude",  "Score 20 posts (hook 1-10, CTA clarity 1-10, pillar) + top 3 gaps",               "_CONTENT_AUDIT/[C]/audit_report.md",          "POSTS_DATA received"),
    ("5", "Content Audit", 3, "Match Winning Templates to ICP",     "🤖 AUTO",    "Claude",  "Rewrite proven CapCut templates for client's niche from swipe file + audit",       "_CONTENT_AUDIT/[C]/template_matches.md",      "audit_report.md done"),

    # DAY 7
    ("7", "Full Content Batch", 1, "Generate 4-Week Content Calendar",   "🤖 AUTO",    "Claude",  "IG + TikTok + Threads + YouTube, 4 weeks",                                        "content_calendar/month_1_calendar.md",        "template_matches.md + swipe_file.md"),
    ("7", "Full Content Batch", 2, "Batch Generate 30 Reel Scripts",     "🤖 AUTO",    "Claude",  "15 Authentic + 15 Value-CTA from matched templates",                              "scripts/reels_batch_1.md",                    "template_matches.md"),
    ("7", "Full Content Batch", 3, "Generate 30 Story Sequences",        "🤖 AUTO",    "Claude",  "14 rotating types: testimonial, value, personal story, win, soft CTA, etc.",      "scripts/stories_batch_1.md",                  "swipe_file.md"),
    ("7", "Full Content Batch", 4, "Generate 30 Threads Posts",          "🤖 AUTO",    "Claude",  "1/day from winning outlier formats",                                              "scripts/threads_batch_1.md",                  "swipe_file.md"),
    ("7", "Full Content Batch", 5, "Generate 4 YouTube Scripts",         "🤖 AUTO",    "Claude",  "From winning outline structures",                                                 "scripts/youtube_batch_1.md",                  "swipe_file.md"),
    ("7", "Full Content Batch", 6, "Generate Voiceover Scripts",         "🤖 AUTO",    "Claude",  "500-word batch, ElevenLabs-ready (breath marks, pacing) for non-talking-head reels","voice_clone/vo_scripts_batch_1.md",          "ELEVENLABS_VOICE_ID collected"),
    ("7", "Full Content Batch", 7, "Generate 30 TikTok Mirror Captions", "🤖 AUTO",    "Claude",  "Rewrite all IG captions for TikTok (shorter, hashtags, CTA style)",               "scripts/tiktok_captions_batch_1.md",          "reels_batch_1.md done"),
    ("7", "Full Content Batch", 8, "Client Calendar Approval",           "📥 COLLECT", "Jordan",  "Share calendar with client → get approval or revision notes",                     "CALENDAR_APPROVED ✓ pasted in worksheet",     "month_1_calendar.md done"),

    # DAY 8
    ("8", "Production + Keyword CTAs", 1, "Convert Scripts → Remotion Props", "🤖 AUTO",    "Claude",  "Generate Remotion JSON files for all 30 reels",                                   "remotion_props/batch_1/",                     "CALENDAR_APPROVED ✓"),
    ("8", "Production + Keyword CTAs", 2, "Generate IG Keyword CTA Pairs",    "🤖 AUTO",    "Claude",  "8-12 keyword + auto-DM reply pairs for Meta Business Suite (free resources only)", "dm_scripts/ig_keyword_cta.md",                "swipe_file.md"),
    ("8", "Production + Keyword CTAs", 3, "Generate TikTok Keyword CTA Setup","🤖 AUTO",    "Claude",  "Keywords → pinned comment text + DM reply templates + linktree structure",         "dm_scripts/tiktok_cta.md",                    "swipe_file.md"),
    ("8", "Production + Keyword CTAs", 4, "CapCut Production",                "🎬 ONSITE",  "Jordan",  "Render Remotion → CapCut polish → sync ElevenLabs VO → dual export IG + TikTok",  "30 reels exported (IG + TikTok versions)",    "Remotion props + Voice ID"),
    ("8", "Production + Keyword CTAs", 5, "Meta Business Suite Keyword CTA",  "⚙️ PLATFORM","Jordan",  "Set up each keyword pair in Meta Business Suite Automations → test each",           "IG_KEYWORD_CTA: live ✓",                      "ig_keyword_cta.md done"),
    ("8", "Production + Keyword CTAs", 6, "TikTok Keyword CTA Setup",         "⚙️ PLATFORM","Jordan",  "Set up Linktree + pin keyword comments + brief setter on TikTok DMs",              "TIKTOK_CTA: live ✓",                          "tiktok_cta.md done"),

    # DAY 9
    ("9", "Repurpose + Ads", 1, "Repurpose 30 Reels → 150+ Pieces",   "🤖 AUTO",    "Claude",  "py -3 tools/repurpose_content.py --batch reels_batch_1.md --client [C]",           "scripts/repurposed_batch_1.md",               "reels_batch_1.md"),
    ("9", "Repurpose + Ads", 2, "Generate All IG Captions + Hashtags", "🤖 AUTO",    "Claude",  "Full captions + hashtag sets for all 30 reels",                                   "scripts/ig_captions_batch_1.md",              "reels_batch_1.md"),
    ("9", "Repurpose + Ads", 3, "Generate Meta Ads Setup + Creatives", "🤖 AUTO",    "Claude",  "Pixel + audience + 3-campaign structure + ad creatives (2 per ad set)",           "analytics/meta_ads_setup.md",                 "offer_sheet.md + swipe_file.md"),
    ("9", "Repurpose + Ads", 4, "Upload TikTok Mirrors",               "⚙️ PLATFORM","Jordan",  "Upload all 30 reels (no watermark) + 30 stories to TikTok",                       "TIKTOK_MIRRORS: uploaded ✓",                  "30 reels exported"),

    # DAY 10
    ("10", "Lead Activation + DFY Assets", 1, "Collect Client's Lead List",           "📥 COLLECT", "Jordan",  "Client exports DM history, email list, old inquiries → shares via GDrive",         "LEAD_LIST: received ✓ + count",               "—"),
    ("10", "Lead Activation + DFY Assets", 2, "Segment Leads (Hot/Warm/Cold)",        "🤖 AUTO",    "Claude",  "Read lead list → segment by engagement recency",                                  "dm_scripts/lead_segments.md",                 "LEAD_LIST received"),
    ("10", "Lead Activation + DFY Assets", 3, "Generate Reactivation Scripts",        "🤖 AUTO",    "Claude",  "DM + Email + WA × Hot/Warm/Cold variants (9 scripts total)",                      "dm_scripts/reactivation_scripts.md",          "lead_segments.md"),
    ("10", "Lead Activation + DFY Assets", 4, "Generate DM Closing Playbook",         "🤖 AUTO",    "Claude",  "Warm-up → opener → qualify → pitch → objection handlers → post-booking",          "dm_scripts/dm_closing_playbook.md",           "swipe_file.md + offer_sheet.md"),
    ("10", "Lead Activation + DFY Assets", 5, "Generate Lead Magnet Content",         "🤖 AUTO",    "Claude",  "Title options + full content + landing page copy + 3 follow-up emails",            "lead_magnet/lead_magnet_content.md",          "swipe_file.md + offer_sheet.md"),
    ("10", "Lead Activation + DFY Assets", 6, "Collect DFY Asset Inputs",             "📥 COLLECT", "Jordan",  "Food prefs, cooking time, family situation, top 5 restaurants, travel frequency",   "Inputs pasted in worksheet",                  "—"),
    ("10", "Lead Activation + DFY Assets", 7, "Generate DFY Asset Pack (Notion-Ready)","🤖 AUTO",   "Claude",  "Shopping list, 7-day meal plan, restaurant guide, travel protocol",                "lead_magnet/dfy_[type].md × 4",               "DFY inputs collected"),
    ("10", "Lead Activation + DFY Assets", 8, "Build Notion Workspace",               "🤖 AUTO",    "Claude",  "Roadmap + bonuses + DFY assets + daily habits + contact table → Notion-ready",    "lead_magnet/notion_workspace_home.md",        "All DFY assets done"),

    # DAY 12
    ("12", "Reactivation + Team SOPs", 1, "Generate Team SOPs (4 Roles)",        "🤖 AUTO",    "Claude",  "Creative Director + Editor + Setter (IG labels + KPI) + Closer",                 "dm_scripts/team_sops.md",                     "swipe_file.md + offer_sheet.md"),
    ("12", "Reactivation + Team SOPs", 2, "Generate Setter DM Scripts (×6 Docs)","🤖 AUTO",    "Claude",  "Cold + warm scraping, sequence notes, DM structure, disqualification, ICP criteria","dm_scripts/setter_scripts_[type].md × 6",     "swipe_file.md + team_sops.md"),
    ("12", "Reactivation + Team SOPs", 3, "First Reactivation Batch",            "🎬 ONSITE",  "Jordan",  "Coach client through sending first 10 reactivation messages + track responses",    "REACTIVATION_RESULTS pasted in worksheet",    "reactivation_scripts.md"),

    # DAY 13
    ("13", "Day 14 Prep", 1, "Generate Day 14 Consultation Brief", "🤖 AUTO",    "Claude",  "Summary of all deliverables + call agenda + goals for months 2-4",               "day14_prep_brief.md",                         "All previous days complete"),
    ("13", "Day 14 Prep", 2, "Pre-Call Verification Checklist",    "🤖 AUTO",    "Claude",  "Check every system is live before the call",                                      "Checklist in chat",                           "day14_prep_brief.md"),

    # DAY 14
    ("14", "Consultation + Launch", 1, "Day 14 Consultation Call",         "🎬 ONSITE",  "Jordan",  "Demo deliverables, launch lead magnet live, test keyword CTA, approve first posts", "FATHOM_CALL link pasted",                     "day14_prep_brief.md"),
    ("14", "Consultation + Launch", 2, "Post-Session Swipe File Update",   "🤖 AUTO",    "Claude",  "Paste call notes → I update swipe_file.md with new context",                      "swipe_file.md updated",                       "Call complete"),

    # DAY 15
    ("15", "Chrome Dashboard + Notion Handover", 1, "Generate Command Center Sheet Template","🤖 AUTO",    "Claude",  "Google Sheets: Daily Dashboard, Content Tracker, Lead Scoreboard, Ad Performance, Weekly KPI, Team Report","analytics/command_center_template.md",       "—"),
    ("15", "Chrome Dashboard + Notion Handover", 2, "Notion Workspace Setup",               "⚙️ PLATFORM","Jordan",  "Paste all Notion-ready files → create subpages → invite client → copy link",        "NOTION_WORKSPACE link pasted",                "notion_workspace_home.md"),
    ("15", "Chrome Dashboard + Notion Handover", 3, "Chrome Dashboard Walkthrough",         "🎬 ONSITE",  "Jordan",  "Screen share Notion, set up Chrome Sales HQ bookmarks, show Google Sheet",          "Sprint complete ✅",                          "Notion live"),
]

# ---------------------------------------------------------------------------
# STYLES
# ---------------------------------------------------------------------------

def hex_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color="D0D0D0")
    return Border(left=s, right=s, top=s, bottom=s)

TAG_COLORS = {
    "🤖 AUTO":    ("1A1A2E", "FFFFFF"),  # dark navy  / white text
    "📥 COLLECT": ("F4A261", "1A1A2E"),  # orange     / dark text
    "🎬 ONSITE":  ("2A9D8F", "FFFFFF"),  # teal       / white text
    "⚙️ PLATFORM":("6A0572", "FFFFFF"),  # purple     / white text
}

WHO_COLORS = {
    "Claude": "E8F4FD",
    "Jordan": "FFF3E0",
    "Client": "E8F5E9",
    "Both":   "F3E5F5",
}

DAY_HEADER_FILL = hex_fill("1A1A2E")
DAY_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)

COL_HEADER_FILL = hex_fill("2D2D2D")
COL_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)

# ---------------------------------------------------------------------------
# BUILD
# ---------------------------------------------------------------------------

def build_sop(ws):
    headers = [
        ("Day",         6),
        ("Phase",       20),
        ("Task #",      7),
        ("Task Name",   32),
        ("Type",        13),
        ("Who",         9),
        ("Description", 55),
        ("Output / File", 38),
        ("Depends On",  30),
        ("Status",      12),
        ("Notes",       25),
    ]

    # Header row
    ws.row_dimensions[1].height = 22
    for col_idx, (label, width) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = COL_HEADER_FILL
        cell.font = COL_HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"

    row = 2
    current_day = None

    for task in TASKS:
        day, phase, task_num, name, tag, who, desc, output, depends = task

        # Day separator row
        if day != current_day:
            current_day = day
            day_label = f"DAY {day} — {phase}"
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
            cell = ws.cell(row=row, column=1, value=day_label)
            cell.fill = DAY_HEADER_FILL
            cell.font = DAY_HEADER_FONT
            cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ws.row_dimensions[row].height = 18
            row += 1

        # Task row
        row_data = [day, phase, task_num, name, tag, who, desc, output, depends, "", ""]
        tag_bg, tag_fg = TAG_COLORS.get(tag, ("FFFFFF", "000000"))
        who_bg = WHO_COLORS.get(who, "FFFFFF")

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin_border()

            # Tag column (col 5): coloured badge
            if col_idx == 5:
                cell.fill = hex_fill(tag_bg)
                cell.font = Font(color=tag_fg, bold=True, size=9)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            # Who column (col 6): coloured by actor
            elif col_idx == 6:
                cell.fill = hex_fill(who_bg)
                cell.font = Font(size=9, bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            # Status column (col 10): dropdown hint
            elif col_idx == 10:
                cell.font = Font(color="888888", italic=True, size=9)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.font = Font(size=9)

        ws.row_dimensions[row].height = 36
        row += 1


def build_legend(ws):
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 48
    ws.column_dimensions["C"].width = 20

    ws.cell(row=1, column=1, value="LEGEND").font = Font(bold=True, size=12)

    tags = [
        ("🤖 AUTO",     "Claude executes immediately — no input needed",        "1A1A2E", "FFFFFF"),
        ("📥 COLLECT",  "Pause here — Jordan or client must provide input",      "F4A261", "1A1A2E"),
        ("🎬 ONSITE",   "Physical action required (filming, call, coaching)",    "2A9D8F", "FFFFFF"),
        ("⚙️ PLATFORM", "UI action in a specific tool (n8n, Netlify, Notion…)",  "6A0572", "FFFFFF"),
    ]

    ws.cell(row=2, column=1, value="Tag").font = Font(bold=True)
    ws.cell(row=2, column=2, value="Meaning").font = Font(bold=True)

    for i, (tag, meaning, bg, fg) in enumerate(tags, start=3):
        c = ws.cell(row=i, column=1, value=tag)
        c.fill = hex_fill(bg)
        c.font = Font(color=fg, bold=True, size=9)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=i, column=2, value=meaning).font = Font(size=9)
        ws.row_dimensions[i].height = 20

    ws.cell(row=8, column=1, value="Who").font = Font(bold=True)
    ws.cell(row=8, column=2, value="Row colour").font = Font(bold=True)

    whos = [
        ("Claude", "E8F4FD", "Blue tint — auto-executed by Claude"),
        ("Jordan", "FFF3E0", "Orange tint — Jordan action required"),
        ("Client", "E8F5E9", "Green tint — client must deliver"),
    ]
    for i, (who, bg, note) in enumerate(whos, start=9):
        c = ws.cell(row=i, column=1, value=who)
        c.fill = hex_fill(bg)
        c.font = Font(bold=True, size=9)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=i, column=2, value=note).font = Font(size=9)
        ws.row_dimensions[i].height = 20

    ws.cell(row=13, column=1, value="Status").font = Font(bold=True)
    for i, s in enumerate(["⬜ Not Started", "🟡 In Progress", "✅ Done"], start=14):
        ws.cell(row=i, column=1, value=s).font = Font(size=9)
        ws.row_dimensions[i].height = 18


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Generate 14-Day Sprint SOP Excel")
    parser.add_argument("--output", default=".tmp/Sprint_SOP.xlsx",
                        help="Output path (default: .tmp/Sprint_SOP.xlsx)")
    args = parser.parse_args()

    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws_sop = wb.active
    ws_sop.title = "Sprint SOP"
    ws_legend = wb.create_sheet("Legend")

    build_sop(ws_sop)
    build_legend(ws_legend)

    wb.save(output_path)
    print(f"[OK]  Sprint SOP saved -> {output_path.resolve()}")


if __name__ == "__main__":
    main()
