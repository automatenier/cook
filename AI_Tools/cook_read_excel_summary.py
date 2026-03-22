import openpyxl
import argparse
import sys
import os
from datetime import datetime

def get_column_indices(sheet):
    """Identify column indices based on header names."""
    headers = [str(cell.value) if cell.value else "" for cell in sheet[1]]
    indices = {}
    
    # Updated mapping to handle numbered headers and specific project patterns
    mapping = {
        'name': ['Name', 'Username', '3-Username', '4-Name', 'BookName'],
        'status': ['Status', '5-Status', 'Stage', '8-Stage', 'close_status'],
        'temp': ['Temperature', '9-Temperature', 'LeadType', '4-leadType'],
        'next_action': ['Next Action', 'Proposed', '27-Proposed'],
        'date': ['Date Added', '13-Leadate', '17-Date', 'date_added'],
        'notes': ['Notes', '14-Notes', '6-Mindmap', '7-Objections', 'objections', 'mindmap']
    }
    
    for key, aliases in mapping.items():
        indices[key] = None # Default to None
        for i, header in enumerate(headers):
            if header and any(alias.lower() in header.lower() for alias in aliases):
                indices[key] = i
                break
    return indices

def summarize_crm(file_path, limit=10):
    if not os.path.exists(file_path):
        return f"Error: File '{file_path}' not found."

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        # Try to find a sensible sheet
        sheet_candidates = ["Leads", "CRM", "Tracker", "Main"]
        sheet = None
        for name in sheet_candidates:
            if name in wb.sheetnames:
                sheet = wb[name]
                break
        if not sheet:
            sheet = wb.active
            
        indices = get_column_indices(sheet)
        
        output = [f"### Summary for {os.path.basename(file_path)}"]
        hot_leads = []
        next_actions = []
        status_counts = {}

        # Safely get values based on indices
        def get_val(row, key):
            idx = indices.get(key)
            if idx is not None and idx < len(row):
                return row[idx]
            return None

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(row): continue
            
            name = get_val(row, 'name') or "Unknown"
            temp = str(get_val(row, 'temp') or "").lower()
            action = get_val(row, 'next_action')
            status = get_val(row, 'status')
            notes = get_val(row, 'notes')

            # Track status counts for the chart
            if status:
                status_str = str(status).strip()
                if status_str:
                    status_counts[status_str] = status_counts.get(status_str, 0) + 1
            
            # Logic for "Hot" - either explicit temperature or specific lead types
            if "hot" in temp or "qualified" in str(status).lower() or "icp" in temp:
                hot_leads.append(f"- 🔥 **{name}**: {status} ({temp})")
            
            if action and action != "None":
                next_actions.append(f"- [ ] **{name}**: {action}")
            elif notes and len(str(notes)) > 5:
                # Fallback to notes if no explicit next action but we have info
                next_actions.append(f"- 📝 **{name}**: {str(notes)[:50]}...")

        # Add Mermaid Chart for Status Distribution
        if status_counts:
            output.append("\n**Lead Status Distribution:**")
            output.append("```mermaid")
            output.append('pie title "Lead Statuses"')
            for s_name, count in status_counts.items():
                output.append(f'    "{s_name}" : {count}')
            output.append("```\n")

        if hot_leads:
            output.append("\n**Key Leads (Qualified/ICP):**")
            output.extend(hot_leads[:limit])
        
        if next_actions:
            output.append("\n**Pending Actions/Notes:**")
            output.extend(next_actions[:limit])
            
        if not hot_leads and not next_actions:
            output.append("\nNo specific leads or pending actions identified.")

        return "\n".join(output)

    except Exception as e:
        return f"Error processing {file_path}: {str(e)}"

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Summarize an Excel CRM file")
    parser.add_argument("file", help="Path to the Excel file")
    parser.add_argument("--limit", type=int, default=10, help="Max items per section")
    args = parser.parse_args()
    
    print(summarize_crm(args.file, args.limit))
