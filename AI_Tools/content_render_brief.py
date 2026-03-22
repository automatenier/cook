"""
Render Brief Generator
======================
Builds a per-client, per-month Excel that serves as the Remotion production spec:
  - Auto-extracts Hook / Value / CTA text from approved .md scripts
  - Leaves Music File, Music Vibe, Overlay Opacity for Jordan to fill
  - Dropdown validation on component, vibe, opacity, status
  - Links to the Obsidian Canvas storyboard column for hard-edit posts

Output:  VLT_Content/02_HMN_HUMANFLOW/jocons/[client]/projects/[YYYY-MM]/render_brief.xlsx
Usage:   py -3 AI_Tools/create_render_brief.py --client fadli --month 2026-03
         py -3 AI_Tools/create_render_brief.py --client fadli          # current month
"""

import re
import sys
import argparse
from pathlib import Path
from datetime import date

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("pip install openpyxl"); sys.exit(1)

ROOT = Path(__file__).parent.parent

# ── Palette (matches other Cook tools) ─────────────────────────────────────────
C = {
    "nav":    "1A1A2E", "card":   "16213E",
    "white":  "FFFFFF", "lgray":  "F4F4F4",
    "mgray":  "AAAAAA", "dgray":  "444444",
    "blue":   "4472C4", "gold":   "FFD700",
    "green":  "C6EFCE", "amber":  "FFD580",
    "red":    "FFDEDE", "teal":   "D6F0EC",
    "purple": "E8D6FF",
}

COMPONENT_MAP = {
    "value + cta": "ValueCTAReel",
    "value+cta":   "ValueCTAReel",
    "value-cta":   "ValueCTAReel",
    "value":       "ValueCTAReel",
    "cta":         "ValueCTAReel",
    "story":       "AuthenticityReel",
    "bts":         "UserVideoReel",
}

STATUS_COLORS = {
    "Pending":      "F4F4F4",
    "Assets Ready": "FFD580",
    "Rendering":    "D6F0EC",
    "Done":         "C6EFCE",
    "Posted":       "E8D6FF",
}

# ── Style helpers ───────────────────────────────────────────────────────────────
def _fill(h): return PatternFill(start_color=h, end_color=h, fill_type="solid")
def _side(c="CCCCCC"): return Side(style="thin", color=c)
def _bdr(c="CCCCCC"): s = _side(c); return Border(left=s, right=s, top=s, bottom=s)
def _font(color="000000", sz=11, bold=False, italic=False):
    return Font(name="Calibri", size=sz, bold=bold, italic=italic, color=color)
def _align(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def cell(ws, row, col, val=None, bg=None, fc="000000", sz=11,
         bold=False, italic=False, h="left", bdr=True, wrap=True):
    c = ws.cell(row=row, column=col)
    if val is not None:
        c.value = val
    if bg:
        c.fill = _fill(bg)
    c.font = _font(fc, sz, bold, italic)
    c.alignment = _align(h, wrap=wrap)
    if bdr:
        c.border = _bdr()
    return c


# ── Script parsing ──────────────────────────────────────────────────────────────
def strip_emotion(text: str) -> str:
    """Remove [emotion] tags like [pensive], [energetic], etc."""
    return re.sub(r"\[.*?\]", "", text).strip()

def strip_markdown(text: str) -> str:
    """Remove bold/italic markdown markers."""
    return re.sub(r"\*+", "", text).strip()

def first_line(text: str) -> str:
    """Return first non-empty line, stripped of emotion tags and markdown."""
    for line in text.splitlines():
        line = strip_emotion(strip_markdown(line)).strip()
        if line:
            return line
    return ""

def extract_cta_keyword(cta_text: str) -> str:
    """Pull the first **bold** word from the CTA section."""
    match = re.search(r"\*\*([A-Z0-9]+)\*\*", cta_text)
    return match.group(1) if match else ""

def parse_script(md_path: Path) -> dict:
    """Parse frontmatter + Hook/Value/CTA sections from an approved script."""
    text = md_path.read_text(encoding="utf-8")

    # Frontmatter
    fm = {}
    fm_match = re.match(r"^---\n(.*?)\n---", text, re.DOTALL)
    if fm_match:
        for line in fm_match.group(1).splitlines():
            if ":" in line:
                k, _, v = line.partition(":")
                fm[k.strip().lower()] = v.strip()

    # Sections
    def get_section(name):
        pattern = rf"#### {name}\s*\n(.*?)(?=\n####|\Z)"
        m = re.search(pattern, text, re.DOTALL | re.IGNORECASE)
        return m.group(1).strip() if m else ""

    hook_raw  = get_section("Hook")
    value_raw = get_section("Value")
    cta_raw   = get_section("CTA")

    content_type = fm.get("type", "").lower().strip()
    component = COMPONENT_MAP.get(content_type, "ValueCTAReel")
    # Also try matching from filename slug pattern
    stem = md_path.stem  # e.g. 2026-03-01_value-cta_diet-consistency
    parts = stem.split("_", 2)
    if len(parts) >= 2 and component == "ValueCTAReel":
        slug_type = parts[1].lower()
        component = COMPONENT_MAP.get(slug_type, component)

    output_name = stem + ".mp4"

    return {
        "date":       fm.get("date", ""),
        "type":       fm.get("type", ""),
        "slug":       fm.get("slug", parts[2] if len(parts) > 2 else stem),
        "component":  component,
        "hook":       first_line(hook_raw),
        "value":      first_line(value_raw),
        "cta":        first_line(cta_raw),
        "cta_keyword": extract_cta_keyword(cta_raw),
        "output":     output_name,
    }


# ── Excel builder ───────────────────────────────────────────────────────────────
HEADERS = [
    "#", "Date", "Type", "Component",
    "Hook Text", "Value Text", "CTA Text", "CTA Keyword",
    "Music File", "Music Vibe", "Volume", "Overlay",
    "Status", "Output File", "Canvas Link", "Notes",
]

COL_WIDTHS = [
    4, 13, 14, 18,
    42, 42, 42, 14,
    22, 16, 9, 10,
    14, 34, 20, 24,
]

def build_sheet(ws, scripts: list, client: str, month: str):
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = C["blue"]
    ws.freeze_panes = "A4"

    # Column widths
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Row heights
    ws.row_dimensions[1].height = 36
    ws.row_dimensions[2].height = 14
    ws.row_dimensions[3].height = 28

    # ── Title row ──
    ws.merge_cells("A1:P1")
    title = ws.cell(row=1, column=1)
    title.value = f"RENDER BRIEF — {client.upper()}  {month}"
    title.font  = Font(name="Calibri", size=16, bold=True, color=C["white"])
    title.fill  = _fill(C["nav"])
    title.alignment = _align("center")

    # ── Spacer row 2 ──
    for col in range(1, len(HEADERS) + 1):
        ws.cell(row=2, column=col).fill = _fill(C["card"])

    # ── Header row 3 ──
    for col, h in enumerate(HEADERS, 1):
        cell(ws, 3, col, h, bg=C["blue"], fc=C["white"], sz=10, bold=True, h="center")

    # ── Data rows ──
    for i, s in enumerate(scripts, 1):
        row = i + 3
        ws.row_dimensions[row].height = 52

        row_bg = C["white"] if i % 2 == 1 else C["lgray"]

        cell(ws, row, 1,  i,                bg=row_bg, h="center", fc=C["mgray"], sz=10)
        cell(ws, row, 2,  s["date"],        bg=row_bg, h="center", bold=True)
        cell(ws, row, 3,  s["type"],        bg=row_bg, h="center", fc="555555", italic=True)
        cell(ws, row, 4,  s["component"],   bg=row_bg, h="center", fc=C["blue"], bold=True)
        cell(ws, row, 5,  s["hook"],        bg=row_bg)
        cell(ws, row, 6,  s["value"],       bg=row_bg, fc="444444")
        cell(ws, row, 7,  s["cta"],         bg=row_bg, fc="444444")
        cell(ws, row, 8,  s["cta_keyword"], bg=row_bg, h="center", bold=True)
        cell(ws, row, 9,  "",               bg="FFF9E6")  # Music File — user fills
        cell(ws, row, 10, "",               bg="FFF9E6")  # Music Vibe — dropdown
        cell(ws, row, 11, 0.25,             bg=row_bg, h="center")
        cell(ws, row, 12, 0.45,             bg=row_bg, h="center")
        cell(ws, row, 13, "Pending",        bg=C["lgray"], h="center")
        cell(ws, row, 14, s["output"],      bg=row_bg, fc="666666", italic=True)
        cell(ws, row, 15, "",               bg=row_bg)  # Canvas Link
        cell(ws, row, 16, "",               bg=row_bg)  # Notes

    # ── Data validation ──
    last_data_row = len(scripts) + 3
    data_range    = f"{last_data_row}"

    # Music Vibe dropdown (col J = 10)
    dv_vibe = DataValidation(
        type="list",
        formula1='"Energetic,Motivational,Chill,Dramatic,Emotional,Hype"',
        allow_blank=True, showDropDown=False,
    )
    dv_vibe.sqref = f"J4:J{last_data_row}"
    ws.add_data_validation(dv_vibe)

    # Overlay dropdown (col L = 12)
    dv_overlay = DataValidation(
        type="list",
        formula1='"0.30,0.45,0.60,0.70"',
        allow_blank=False, showDropDown=False,
    )
    dv_overlay.sqref = f"L4:L{last_data_row}"
    ws.add_data_validation(dv_overlay)

    # Component dropdown (col D = 4)
    dv_comp = DataValidation(
        type="list",
        formula1='"ValueCTAReel,AuthenticityReel,UserVideoReel"',
        allow_blank=False, showDropDown=False,
    )
    dv_comp.sqref = f"D4:D{last_data_row}"
    ws.add_data_validation(dv_comp)

    # Status dropdown (col M = 13)
    dv_status = DataValidation(
        type="list",
        formula1='"Pending,Assets Ready,Rendering,Done,Posted"',
        allow_blank=False, showDropDown=False,
    )
    dv_status.sqref = f"M4:M{last_data_row}"
    ws.add_data_validation(dv_status)

    # Conditional fill for Status column — color-code by value
    from openpyxl.formatting.rule import CellIsRule
    for status, color in STATUS_COLORS.items():
        rule = CellIsRule(
            operator="equal",
            formula=[f'"{status}"'],
            fill=_fill(color),
            font=Font(name="Calibri", size=11, color="000000"),
        )
        ws.conditional_formatting.add(f"M4:M{last_data_row}", rule)

    # ── Legend row ──
    legend_row = last_data_row + 2
    ws.merge_cells(f"A{legend_row}:D{legend_row}")
    leg = ws.cell(row=legend_row, column=1)
    leg.value = "Yellow cells = fill before rendering   |   Music vol default 0.25   |   Overlay 0.45 = standard"
    leg.font  = Font(name="Calibri", size=9, italic=True, color=C["mgray"])
    leg.alignment = _align("left", wrap=False)


def build_instructions(ws):
    """Quick-ref tab explaining how to use the brief."""
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = C["gold"][:-2] + "00"
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 60

    ws.merge_cells("A1:C1")
    t = ws.cell(row=1, column=1)
    t.value = "HOW TO USE THIS RENDER BRIEF"
    t.font  = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    t.fill  = _fill(C["nav"])
    t.alignment = _align("center")

    steps = [
        ("Step 1", "Fill Music File (col I)",
         "Drop your music file into VLT_Content/AI_ENGINE/remotion/public/ as music.mp3\n"
         "Then write the filename here (e.g. hype_track_01.mp3)"),
        ("Step 2", "Pick Music Vibe (col J)",
         "Choose from dropdown — this is just a reference label for you, not used by Remotion"),
        ("Step 3", "Set Overlay (col L)",
         "0.30 = footage is dark / high contrast\n0.45 = standard (default)\n0.60 = bright footage, need darker overlay for text"),
        ("Step 4", "Canvas storyboard (col O)",
         "Paste the Obsidian Canvas filename here for hard-edit posts.\n"
         "Canvas = which clips go where (footage order).\nThis Excel = Remotion settings (text, music, overlay)."),
        ("Step 5", "Update defaultProps (Root.tsx)",
         "Copy Hook / Value / CTA text from this sheet into\n"
         "VLT_Content/AI_ENGINE/remotion/src/Root.tsx → defaultProps\nbefore running the render."),
        ("Step 6", "Render",
         "Drop user-video.mp4 + music.mp3 into remotion/public/\n"
         "Then: cd VLT_Content/AI_ENGINE/remotion && npm run build\n"
         "Or batch: py -3 AI_Tools/render_batch_modal.py --client [name]"),
        ("Step 7", "Update Status",
         "Pending → Assets Ready → Rendering → Done → Posted\nStatus column auto-colors."),
    ]

    for i, (step, title, desc) in enumerate(steps, 3):
        ws.row_dimensions[i].height = 56
        cell(ws, i, 2, f"{step}: {title}", bg="F0F4FF", bold=True, sz=11)
        cell(ws, i, 3, desc, bg=C["white"], sz=10, italic=True)


# ── Entry point ─────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Generate Remotion render brief Excel")
    parser.add_argument("--client", required=True, help="Client name (e.g. fadli)")
    parser.add_argument("--month",  default=date.today().strftime("%Y-%m"),
                        help="Month in YYYY-MM format (default: current month)")
    args = parser.parse_args()

    client = args.client.lower()
    month  = args.month

    approved_dir = ROOT / "Content" / "02_WORKSPACE" / "jocons" / client / "projects" / month / "approved"
    if not approved_dir.exists():
        print(f"ERROR: Approved scripts not found at {approved_dir}")
        sys.exit(1)

    md_files = sorted(f for f in approved_dir.glob("*.md")
                      if re.match(r"\d{4}-\d{2}-\d{2}", f.name))
    if not md_files:
        print(f"ERROR: No .md scripts found in {approved_dir}")
        sys.exit(1)

    print(f"Found {len(md_files)} scripts for {client} / {month}")
    scripts = []
    for f in md_files:
        try:
            scripts.append(parse_script(f))
            print(f"  OK {f.name}")
        except Exception as e:
            print(f"  FAIL {f.name} - {e}")

    wb = Workbook()
    ws_brief = wb.active
    ws_brief.title = "Render Brief"
    ws_how   = wb.create_sheet("How To Use")

    build_sheet(ws_brief, scripts, client, month)
    build_instructions(ws_how)

    out_path = approved_dir.parent / "render_brief.xlsx"
    wb.save(out_path)
    print(f"\nSaved -> {out_path}")


if __name__ == "__main__":
    main()
