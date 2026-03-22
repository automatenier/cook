import sys, io, shutil
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

from pathlib import Path

ROOT = Path(__file__).parent.parent
SRC  = ROOT / '01 HMN__Command/00_COOK/Daily_KPI_Tracker.xlsx'
BACK = ROOT / '.tmp/Daily_KPI_Tracker_backup.xlsx'

shutil.copy2(SRC, BACK)
print("Backup saved.")

wb = openpyxl.load_workbook(SRC)

# ── Colours ───────────────────────────────────────────────────────────────────
JO_BG   = "2E4057"   # dark navy  – JO Consult headers
RE_BG   = "048A81"   # teal       – Real Estate headers
HDR_BG  = "1A1A2E"   # near-black – main headers
JO_ROW  = "EEF2FF"   # light lavender  – JO alternating row
RE_ROW  = "E6F7F5"   # light mint      – RE alternating row
WHITE   = "FFFFFF"
GREY    = "F2F2F2"

# ── Helper: style a single cell ───────────────────────────────────────────────
def style(cell, value="", bold=False, fg=WHITE, bg=None,
          size=10, halign="center", wrap=False, border=False):
    cell.value = value
    cell.font  = Font(bold=bold, color=fg, size=size)
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=halign, vertical="center", wrap_text=wrap)
    if border:
        thin = Side(style="thin", color="CCCCCC")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return cell

# ── 1. ADD VERTICAL COLUMN to Content, Setter, Sales ─────────────────────────
TABS = {
    "Content": {"last_col": 11, "task_cols": "C:I", "n_tasks": 7},
    "Setter":  {"last_col": 10, "task_cols": "C:H", "n_tasks": 6},
    "Sales":   {"last_col": 10, "task_cols": "C:H", "n_tasks": 6},
}

for sheet_name, info in TABS.items():
    # Find the actual sheet (emoji prefix possible)
    ws = next((wb[s] for s in wb.sheetnames if sheet_name in s), None)
    if ws is None:
        print(f"  WARN: sheet '{sheet_name}' not found, skipping")
        continue

    new_col = info["last_col"] + 1
    col_ltr = get_column_letter(new_col)

    # Header cell (row 3)
    hdr_cell = ws.cell(row=3, column=new_col)
    style(hdr_cell, "Vertical", bold=True, fg=WHITE, bg=HDR_BG, border=True)
    ws.column_dimensions[col_ltr].width = 14

    # Data validation dropdown for rows 4 → max_row
    dv = DataValidation(
        type="list",
        formula1='"JO Consult,Real Estate"',
        allow_blank=True,
        showDropDown=False,
    )
    dv.sqref = f"{col_ltr}4:{col_ltr}{ws.max_row}"
    ws.add_data_validation(dv)

    # Light styling for data cells
    for r in range(4, ws.max_row + 1):
        c = ws.cell(row=r, column=new_col)
        bg = JO_ROW if r % 2 == 0 else WHITE
        style(c, bg=bg, border=True)

    print(f"  {sheet_name}: Vertical column added at col {col_ltr}")

# ── 2. CREATE 📊 Vertical Summary TAB (inserted as tab 0) ─────────────────────
if "📊 Vertical Summary" in wb.sheetnames:
    del wb["📊 Vertical Summary"]

vs = wb.create_sheet("📊 Vertical Summary", 0)
vs.sheet_view.showGridLines = False
vs.sheet_view.zoomScale = 100

# Column widths
col_widths = {
    "A": 30, "B": 13, "C": 13, "D": 13, "E": 16, "F": 10,
    "G": 3,
    "H": 30, "I": 13, "J": 13, "K": 13, "L": 16, "M": 10,
}
for col, w in col_widths.items():
    vs.column_dimensions[col].width = w

# Row heights helper
def rh(row, h): vs.row_dimensions[row].height = h

# ── Title ──────────────────────────────────────────────────────────────────────
vs.merge_cells("A1:M1")
style(vs["A1"], "📊  VERTICAL KPI SUMMARY — JO Consult & Real Estate",
      bold=True, size=14, bg=HDR_BG, fg=WHITE)
rh(1, 36)
rh(2, 8)   # spacer

# ── Section headers row 3 ──────────────────────────────────────────────────────
vs.merge_cells("A3:F3")
style(vs["A3"], "🏢  JO CONSULT", bold=True, size=12, bg=JO_BG, fg=WHITE)
vs.merge_cells("H3:M3")
style(vs["H3"], "🏡  REAL ESTATE", bold=True, size=12, bg=RE_BG, fg=WHITE)
rh(3, 28)

# ── Sub-headers row 4 ─────────────────────────────────────────────────────────
JO_SUB = ["KPI", "Today", "This Week", "This Month", "Target", "Status"]
RE_SUB = ["KPI", "Today", "This Week", "This Month", "Target", "Status"]

for i, h in enumerate(JO_SUB):
    style(vs.cell(4, 1 + i), h, bold=True, bg=JO_BG, fg=WHITE, border=True, size=9)
for i, h in enumerate(RE_SUB):
    style(vs.cell(4, 8 + i), h, bold=True, bg=RE_BG, fg=WHITE, border=True, size=9)
rh(4, 20)

# ── KPI rows ──────────────────────────────────────────────────────────────────
JO_KPIS = [
    ("📩 DMs Sent Today",            "30+"),
    ("🔄 Follow Ups Sent",           "10"),
    ("📱 Reels Posted",              "1/day"),
    ("✍️ Scripts Written",           "1/day"),
    ("📞 Discovery Calls Booked",    "≥1"),
    ("📋 Proposals Sent",            "count"),
    ("🔥 Hot Leads in Pipeline",     "count"),
    ("💰 Cash Collected (IDR)",      "manual"),
]

RE_KPIS = [
    ("📩 DMs Sent Today",            "20+"),
    ("🔄 Follow Ups Sent",           "10"),
    ("📱 Reels Posted",              "1/day"),
    ("🏠 Viewings Booked",           "count"),
    ("🏠 Viewings Conducted",        "count"),
    ("🤝 Deals in Negotiation",      "count"),
    ("📋 Active Listings",           "count"),
    ("💰 Commission Pipeline (IDR)", "manual"),
]

for i, (kpi, target) in enumerate(JO_KPIS):
    r = 5 + i
    bg = JO_ROW if i % 2 == 0 else WHITE
    style(vs.cell(r, 1), kpi,    bold=True, bg=bg, fg="1A1A2E", halign="left",  border=True)
    style(vs.cell(r, 2), "",     bg=bg, border=True)
    style(vs.cell(r, 3), "",     bg=bg, border=True)
    style(vs.cell(r, 4), "",     bg=bg, border=True)
    style(vs.cell(r, 5), target, bg=bg, fg="555555", border=True, size=9)
    style(vs.cell(r, 6), "",     bg=bg, border=True)
    rh(r, 20)

for i, (kpi, target) in enumerate(RE_KPIS):
    r = 5 + i
    bg = RE_ROW if i % 2 == 0 else WHITE
    style(vs.cell(r, 8),  kpi,    bold=True, bg=bg, fg="1A1A2E", halign="left", border=True)
    style(vs.cell(r, 9),  "",     bg=bg, border=True)
    style(vs.cell(r, 10), "",     bg=bg, border=True)
    style(vs.cell(r, 11), "",     bg=bg, border=True)
    style(vs.cell(r, 12), target, bg=bg, fg="555555", border=True, size=9)
    style(vs.cell(r, 13), "",     bg=bg, border=True)

last_kpi_row = 5 + max(len(JO_KPIS), len(RE_KPIS))
rh(last_kpi_row + 1, 14)  # spacer

# ── Monthly Revenue section ────────────────────────────────────────────────────
rev_r = last_kpi_row + 2

vs.merge_cells(f"A{rev_r}:F{rev_r}")
style(vs[f"A{rev_r}"], "💰  MONTHLY REVENUE — JO CONSULT (IDR)",
      bold=True, size=11, bg=JO_BG, fg=WHITE)
vs.merge_cells(f"H{rev_r}:M{rev_r}")
style(vs[f"H{rev_r}"], "💰  MONTHLY COMMISSION — REAL ESTATE (IDR)",
      bold=True, size=11, bg=RE_BG, fg=WHITE)
rh(rev_r, 24)

MONTHS_H1 = ["Jan", "Feb", "Mar", "Apr", "May", "Jun"]
MONTHS_H2 = ["Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

for i, m in enumerate(MONTHS_H1):
    col_jo = 1 + i
    col_re = 8 + i
    style(vs.cell(rev_r+1, col_jo), m, bold=True, bg="E8F0FE", fg=HDR_BG, border=True)
    style(vs.cell(rev_r+2, col_jo), "", bg=WHITE, border=True)
    style(vs.cell(rev_r+1, col_re), m, bold=True, bg="E8F9F7", fg=HDR_BG, border=True)
    style(vs.cell(rev_r+2, col_re), "", bg=WHITE, border=True)

for i, m in enumerate(MONTHS_H2):
    col_jo = 1 + i
    col_re = 8 + i
    style(vs.cell(rev_r+3, col_jo), m, bold=True, bg="E8F0FE", fg=HDR_BG, border=True)
    style(vs.cell(rev_r+4, col_jo), "", bg=WHITE, border=True)
    style(vs.cell(rev_r+3, col_re), m, bold=True, bg="E8F9F7", fg=HDR_BG, border=True)
    style(vs.cell(rev_r+4, col_re), "", bg=WHITE, border=True)

# Annual totals
style(vs.cell(rev_r+5, 1), "Annual Total (IDR)", bold=True, bg=GREY, fg=HDR_BG,
      halign="left", border=True)
style(vs.cell(rev_r+5, 8), "Annual Total (IDR)", bold=True, bg=GREY, fg=HDR_BG,
      halign="left", border=True)
for c in range(2, 7):
    style(vs.cell(rev_r+5, c), "", bg=GREY, border=True)
for c in range(9, 14):
    style(vs.cell(rev_r+5, c), "", bg=GREY, border=True)

for r_off in range(1, 6):
    rh(rev_r + r_off, 20)

# ── Instructions footer ────────────────────────────────────────────────────────
foot_r = rev_r + 7
vs.merge_cells(f"A{foot_r}:M{foot_r}")
style(vs[f"A{foot_r}"],
      "📝  Fill Today / This Week / This Month manually each morning. "
      "Use the Vertical dropdown in Content, Setter, Sales tabs to tag each row.",
      fg="555555", size=9, bg="FAFAFA", halign="left", wrap=True)
rh(foot_r, 28)

# ── Save ──────────────────────────────────────────────────────────────────────
wb.save(SRC)
print(f"Saved → {SRC}")
print("Done.")
