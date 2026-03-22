import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PATH = r"c:\Users\natha\OneDrive - Bina Nusantara\Cook\HMN_A Human\00_AGENT_COMMAND.xlsx"

# ── Colors ────────────────────────────────────────────────────────────────────
C_HEADER_BG  = "1A1A2E"   # dark navy
C_HEADER_FG  = "FFFFFF"
C_SEC_BG     = "16213E"   # section title bg
C_SEC_FG     = "E2B96E"   # gold
C_GREEN      = "D6F4D1"   # working
C_YELLOW     = "FFF3CC"   # minimal / partial
C_RED        = "FFD6D6"   # to build
C_BLUE       = "D6E8FF"   # new / imported
C_ROW_ALT    = "F7F9FC"   # alt row stripe
C_BORDER     = "CCCCCC"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def bold(color="000000", size=10):
    return Font(bold=True, color=color, size=size)

def thin_border():
    s = Side(style="thin", color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

# ── Data ──────────────────────────────────────────────────────────────────────
# Columns: Domain | Automation | Trigger | Status | Roadmap Phase | Priority | Notes
AUTOMATIONS = [
    # AGENCY OPS
    ("Agency OPS", "Calendar-Assign-Setup",      "Google Calendar event",    "✅ Working",  "Phase 3: Booking calendar",         "HIGH",   ""),
    ("Agency OPS", "Form-1-Onboarding",          "Form submission",          "✅ Working",  "Phase 1: Onboarding",               "HIGH",   ""),
    ("Agency OPS", "Form-B-Invoice",             "Form submission",          "✅ Working",  "Phase 0: Payment/invoicing",        "HIGH",   ""),
    ("Agency OPS", "Form-A-Legal",               "Form submission",          "⚠️ Minimal",  "—",                                 "LOW",    "Empty/minimal content"),
    ("Agency OPS", "Form-A-Privacy",             "Form submission",          "⚠️ Minimal",  "—",                                 "LOW",    "Empty/minimal content"),
    ("Agency OPS", "Form-B-Add-ons",             "Form submission",          "⚠️ Minimal",  "—",                                 "LOW",    ""),
    ("Agency OPS", "Form-CONS-Content",          "Form submission",          "✅ Working",  "Phase 4: Content audit",            "HIGH",   ""),
    ("Agency OPS", "Form-CONS-Leads",            "Form submission",          "✅ Working",  "Phase 7: Lead activation",          "HIGH",   ""),
    ("Agency OPS", "Form-CONS-Program",          "Form submission",          "✅ Working",  "Phase 8: Consultation",             "HIGH",   ""),
    ("Agency OPS", "Form-CONS-Technics",         "Form submission",          "⚠️ Minimal",  "—",                                 "LOW",    ""),
    ("Agency OPS", "Form-StratPDF",              "Form submission",          "✅ Working",  "Phase 2: Strategy output",          "HIGH",   ""),
    ("Agency OPS", "Form-AuditContent",          "Scheduled / webhook",      "✅ Working",  "Phase 4: Content monitoring",       "HIGH",   ""),
    ("Agency OPS", "Notif-Error",                "Workflow error trigger",   "✅ Working",  "All phases: Error alerting",        "HIGH",   "Monitors all workflows"),
    ("Agency OPS", "Agency-OPS-BOT",             "Telegram command",         "✅ Working",  "Internal: PDF gen + management",   "HIGH",   ""),

    # AGENCY ACQ
    ("Agency ACQ", "Chatbot-JO (WhatsApp)",      "WhatsApp message",         "✅ Working",  "Outreach: Lead qualification",      "HIGH",   ""),
    ("Agency ACQ", "DND-Reel Jo",                "Telegram command",         "✅ Working",  "Phase 5: Content downloading",      "HIGH",   "Jordan's reel downloader"),
    ("Agency ACQ", "DND-Reel Man",               "Telegram command",         "✅ Working",  "Phase 5: Content downloading",      "MED",    "Male reference reels"),
    ("Agency ACQ", "DND-Reel Woman",             "Telegram command",         "✅ Working",  "Phase 5: Content downloading",      "MED",    "Female / Gemini analysis"),
    ("Agency ACQ", "SaveImage",                  "Telegram image receive",   "✅ Working",  "Utility: Asset storage",            "MED",    "Telegram → GDrive"),
    ("Agency ACQ", "GMAPS Scraper",              "Manual / webhook",         "✅ Working",  "Outreach: Lead generation",         "HIGH",   ""),
    ("Agency ACQ", "Gmaps (variant)",            "Manual / webhook",         "✅ Working",  "Outreach: Lead generation",         "MED",    "Alt variant"),
    ("Agency ACQ", "IG Comments",                "Apify webhook",            "✅ Working",  "Outreach: IG engagement analysis",  "HIGH",   "Claude analysis included"),
    ("Agency ACQ", "IG DMS",                     "Manual / scheduled",       "✅ Working",  "Outreach: Automated DM",            "HIGH",   "Puppeteer-based"),
    ("Agency ACQ", "IG Profile",                 "Manual / webhook",         "✅ Working",  "Outreach: Profile scraping",        "MED",    ""),
    ("Agency ACQ", "Linkedin Jobs",              "Manual / webhook",         "✅ Working",  "Outreach: Job/lead scraping",       "MED",    ""),
    ("Agency ACQ", "Youtube Scraper",            "Manual / webhook",         "✅ Working",  "Content: YouTube analysis",         "MED",    ""),
    ("Agency ACQ", "Notif-NOSHOW-WA",            "Webhook / manual",         "✅ Working",  "Phase 7-9: No-show WA",             "HIGH",   ""),
    ("Agency ACQ", "Notif-Sales Call",           "Google Calendar",          "✅ Working",  "Phase 7-9: Call reminders (TG)",    "HIGH",   ""),
    ("Agency ACQ", "Notif-SalesCall-WA Reminder","Google Calendar",          "✅ Working",  "Phase 7-9: WA reminder 24h",        "HIGH",   "Partial — may not cover full Wassenger flow"),
    ("Agency ACQ", "VSL-Chatbot-Jule",           "Website chat widget",      "🆕 New",      "Outreach: VSL page chatbot",        "HIGH",   "Needs import/activate"),
    ("Agency ACQ", "ElevenLabs-Voice-Call",      "Webhook",                  "🆕 New",      "Outreach: AI voice call agent",     "HIGH",   "Needs import/activate"),

    # CLIENTS
    ("Clients",    "Reel Analyzer",              "Telegram command",         "✅ Working",  "Phase 5: Content analysis",         "HIGH",   "Universal — all clients"),
    ("Clients",    "ACC-BOT",                    "Telegram command",         "✅ Working",  "Client: Accountability bot",        "HIGH",   ""),
    ("Clients",    "Coach-BOT",                  "Telegram command",         "✅ Working",  "Client: Coach management",          "HIGH",   ""),
    ("Clients",    "Sales-BOT",                  "Telegram command",         "✅ Working",  "Client: Sales tracking",            "HIGH",   ""),
    ("Clients",    "TRAINEE-BOT",                "Telegram command",         "✅ Working",  "Client: Trainee tracking",          "HIGH",   ""),
    ("Clients",    "Life Story Form",            "Form submission",          "✅ Working",  "Phase 1.5: Client info collection", "HIGH",   ""),
    ("Clients",    "Notif-1stCoachCall",         "Google Calendar",          "✅ Working",  "Phase 8: First consultation",       "HIGH",   ""),
    ("Clients",    "Notif-Setup",                "Webhook / trigger",        "✅ Working",  "Phase 0-1: Setup notifications",    "HIGH",   ""),
    ("Clients",    "Notif-Tugas",                "Scheduled / webhook",      "✅ Working",  "Ongoing: Task notifications",       "HIGH",   ""),
    ("Clients",    "Notif-DLY-Workout",          "Cron daily",               "✅ Working",  "Client product: Daily workout",     "HIGH",   ""),
    ("Clients",    "Post-Booking Elevenlabs",    "Google Calendar booking",  "✅ Working",  "Phase 1.3: Voice clone / TTS",      "HIGH",   ""),
    ("Clients",    "EOD-DLY",                    "Cron daily (EOD)",         "✅ Working",  "Phase 9: Daily setter report",      "HIGH",   ""),
    ("Clients",    "EOD-WKLY",                   "Cron weekly",              "✅ Working",  "Phase 9: Weekly report",            "HIGH",   ""),
    ("Clients",    "Notif-Close-Invoice",        "Webhook / manual",         "✅ Working",  "Phase 0: Close + invoice",          "HIGH",   ""),
    ("Clients",    "Notif-DLY-KPI",             "Cron daily",               "✅ Working",  "Phase 9: KPI tracking",             "HIGH",   ""),
    ("Clients",    "Notif-Posts",               "Webhook (post published)",  "✅ Working",  "Phase 10: Post notifications",      "MED",    "Sales-facing, not client-facing"),
    ("Clients",    "Notif-Setter",              "Webhook / scheduled",       "✅ Working",  "Phase 9: Setter notifications",     "HIGH",   ""),
    ("Clients",    "Report-MO",                 "Cron monthly",              "✅ Working",  "Ongoing: Monthly report",           "HIGH",   ""),
    ("Clients",    "Form-Meal",                 "Form submission",           "✅ Working",  "Client product: Meal tracking",     "HIGH",   ""),

    # JORDAN TOOLS
    ("Jordan",     "Agency-OPS-BOT",            "Telegram command",          "✅ Working",  "Internal: OPS + PDF gen",           "HIGH",   ""),
    ("Jordan",     "WhatsApp-ElevenLabs-Agent", "WhatsApp message",          "🆕 Draft",    "Internal: Personal WA voice agent", "MED",    "Draft — not deployed"),
    ("Jordan",     "Notif-Error",               "Workflow error",            "✅ Working",  "All workflows: Error alerting",     "HIGH",   ""),
]

GAPS = [
    # (Priority, #, Name, Phase, Description, Build Order)
    ("HIGH",   1, "Telegram Daily Client Check-in",    "Phase 11",   "Cron 08:00 WIB → pull today tasks → Claude formats → send to client TG channel",            "Build NOW"),
    ("HIGH",   2, "Telegram Weekly Progress Summary",  "Phase 11",   "Cron Sunday 20:00 WIB → pull week metrics → Claude summarizes → scorecard to client TG",     "Build NOW"),
    ("HIGH",   3, "Telegram Team Daily Briefing",      "Phase 11",   "Cron 08:30 WIB → all clients' tasks → per-client brief → internal team TG channel",          "Build NOW"),
    ("MED",    4, "Telegram Personal Dashboard",       "Phase 11",   "Cron Monday 07:00 WIB → all-client metrics + revenue + pipeline → Jordan's personal TG",     "Build NOW"),
    ("HIGH",   5, "WA Post-Booking Reminder (24h)",    "Phase 3.3",  "GCal booking → Wassenger WA reminder 24h before with prep tips",                            "Build NEXT"),
    ("HIGH",   6, "Email Post-Booking Reminder (1h)",  "Phase 3.3",  "GCal event 1h before → email with Zoom link + agenda",                                       "Build NEXT"),
    ("MED",    7, "No-Show Auto Follow-Up (Multi-Ch)", "Phase 3.3",  "No-show detected → WA → wait 2h → Email → wait 24h → DM (via IG DMS)",                     "Build NEXT"),
    ("MED",    8, "Content Published Notification",    "Phase 11",   "Content goes live → notify client via TG with link + metrics preview",                       "Build NEXT"),
    ("LOW",    9, "Lead Magnet Download Counter",      "Phase 11",   "Track lead magnet downloads → update client TG",                                             "Nice-to-have"),
    ("LOW",   10, "Milestone Celebrations",            "Phase 11",   "Auto-detect milestones (10 clients, 100 leads, 1000 followers) → celebration message",       "Nice-to-have"),
    ("MED",   11, "Client Folder Auto-Creator",        "Phase 0",    "Payment confirmed → auto-create full GDrive folder structure",                               "Nice-to-have"),
    ("MED",   12, "Onboarding Progress Tracker",       "Phase 1",    "Track which onboarding steps completed per client → alert on incomplete",                   "Nice-to-have"),
    ("LOW",   13, "Content Calendar Auto-Publisher",   "Phase 10",   "Pull approved content → auto-schedule in Meta Biz Suite / TikTok",                          "Nice-to-have"),
    ("LOW",   14, "Reactivation Campaign Tracker",     "Phase 7",    "Track reactivation messages sent/replied/booked per segment",                               "Nice-to-have"),
    ("LOW",   15, "Ad Spend vs Leads Webhook",         "Phase 6",    "Pull Meta Ads data → compare to leads generated → weekly summary",                          "Nice-to-have"),
]

# ── Build sheet ───────────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(PATH)

SHEET_NAME = "🔧 n8n Workflows"
if SHEET_NAME in wb.sheetnames:
    del wb[SHEET_NAME]
ws = wb.create_sheet(SHEET_NAME)

row = 1

def write_title(ws, row, text):
    ws.merge_cells(f"A{row}:G{row}")
    c = ws.cell(row, 1, text)
    c.fill = fill(C_HEADER_BG)
    c.font = Font(bold=True, color=C_HEADER_FG, size=13)
    c.alignment = center()
    ws.row_dimensions[row].height = 24
    return row + 1

def write_col_headers(ws, row, headers):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row, col, h)
        c.fill = fill("2E4057")
        c.font = Font(bold=True, color="FFFFFF", size=9)
        c.alignment = center()
        c.border = thin_border()
    ws.row_dimensions[row].height = 20
    return row + 1

def write_section(ws, row, label):
    ws.merge_cells(f"A{row}:G{row}")
    c = ws.cell(row, 1, label)
    c.fill = fill(C_SEC_BG)
    c.font = Font(bold=True, color=C_SEC_FG, size=9)
    c.alignment = left()
    ws.row_dimensions[row].height = 16
    return row + 1

STATUS_COLOR = {
    "✅": C_GREEN,
    "⚠️": C_YELLOW,
    "❌": C_RED,
    "🆕": C_BLUE,
}

def status_fill(status):
    for k, v in STATUS_COLOR.items():
        if k in status:
            return fill(v)
    return fill("FFFFFF")

# ── TITLE ─────────────────────────────────────────────────────────────────────
row = write_title(ws, row, "N8N WORKFLOW MONITOR — Cook Automation Stack")

# blank row
row += 1

# ── SECTION 1: Existing Automations ──────────────────────────────────────────
ws.merge_cells(f"A{row}:G{row}")
c = ws.cell(row, 1, f"EXISTING AUTOMATIONS  ({len(AUTOMATIONS)} workflows)")
c.fill = fill("0F3460")
c.font = Font(bold=True, color="E2B96E", size=10)
c.alignment = left()
row += 1

HEADERS = ["Domain", "Automation Name", "Trigger", "Status", "Roadmap Phase", "Priority", "Notes"]
row = write_col_headers(ws, row, HEADERS)

current_domain = None
for i, (domain, name, trigger, status, phase, priority, notes) in enumerate(AUTOMATIONS):
    if domain != current_domain:
        row = write_section(ws, row, f"  {domain}")
        current_domain = domain

    bg = C_ROW_ALT if i % 2 == 0 else "FFFFFF"
    values = [domain, name, trigger, status, phase, priority, notes]
    for col, val in enumerate(values, 1):
        c = ws.cell(row, col, val)
        c.border = thin_border()
        c.alignment = left()
        c.font = Font(size=9)
        if col == 4:  # Status column
            c.fill = status_fill(status)
            c.alignment = center()
        elif col == 6:  # Priority
            c.alignment = center()
            pf = {"HIGH": fill("FFD6D6"), "MED": fill("FFF3CC"), "LOW": fill(C_ROW_ALT)}
            c.fill = pf.get(priority, fill("FFFFFF"))
        else:
            c.fill = fill(bg)
    ws.row_dimensions[row].height = 15
    row += 1

# blank row
row += 1

# ── SECTION 2: Gap List ───────────────────────────────────────────────────────
ws.merge_cells(f"A{row}:G{row}")
c = ws.cell(row, 1, f"GAP LIST — MISSING AUTOMATIONS  ({len(GAPS)} workflows to build)")
c.fill = fill("0F3460")
c.font = Font(bold=True, color="FFB3B3", size=10)
c.alignment = left()
row += 1

GAP_HEADERS = ["Priority", "#", "Automation Name", "Roadmap Phase", "Description", "Build Order", "Status"]
row = write_col_headers(ws, row, GAP_HEADERS)

for i, (priority, num, name, phase, desc, build_order) in enumerate(GAPS):
    bg = C_ROW_ALT if i % 2 == 0 else "FFFFFF"
    values = [priority, num, name, phase, desc, build_order, "❌ Not Built"]
    for col, val in enumerate(values, 1):
        c = ws.cell(row, col, val)
        c.border = thin_border()
        c.alignment = left()
        c.font = Font(size=9)
        if col == 1:  # Priority
            c.alignment = center()
            pf = {"HIGH": fill("FFD6D6"), "MED": fill("FFF3CC"), "LOW": fill(C_ROW_ALT)}
            c.fill = pf.get(priority, fill("FFFFFF"))
        elif col == 7:  # Status
            c.fill = fill(C_RED)
            c.alignment = center()
        elif col == 6:  # Build order
            bo_colors = {"Build NOW": fill("FFD6D6"), "Build NEXT": fill("FFF3CC"), "Nice-to-have": fill(C_ROW_ALT)}
            c.fill = bo_colors.get(build_order, fill("FFFFFF"))
            c.alignment = center()
        else:
            c.fill = fill(bg)
    ws.row_dimensions[row].height = 28
    row += 1

# blank row
row += 1

# ── SECTION 3: Coverage Summary ───────────────────────────────────────────────
ws.merge_cells(f"A{row}:G{row}")
c = ws.cell(row, 1, "ROADMAP COVERAGE SUMMARY")
c.fill = fill("0F3460")
c.font = Font(bold=True, color="E2B96E", size=10)
c.alignment = left()
row += 1

COV_HEADERS = ["Phase", "Coverage", "Gap Level", "", "", "", ""]
row = write_col_headers(ws, row, COV_HEADERS)

COVERAGE = [
    ("Phase 0: Payment + CRM",      "✅ Invoice, Close notifications",         "Minor: Auto folder creation",     ""),
    ("Phase 1: Onboarding",          "✅ Forms, Life Story, Setup notifs",       "Minor: Progress tracking",        ""),
    ("Phase 2: Offer Blueprint",     "✅ Strategy PDF, forms",                   "Covered",                         ""),
    ("Phase 3: Funnel + Calendar",   "✅ Calendar assign + WA reminder",         "GAP: Email 1h reminder",          "HIGH"),
    ("Phase 4: Content Audit",       "✅ Audit forms, monitoring",               "Covered",                         ""),
    ("Phase 5: Content Production",  "✅ Reel analysis, downloading",            "Covered",                         ""),
    ("Phase 6: CTA + Ads",           "⚠️ No automation found",                  "Low priority (manual setup)",     "LOW"),
    ("Phase 7: Lead Activation",     "✅ Lead forms",                            "Minor: Tracker",                  ""),
    ("Phase 8: Consultation",        "✅ 1st coach call, program form",          "Covered",                         ""),
    ("Phase 9: Team SOPs + Reports", "✅ EOD daily/weekly, KPI, setter notifs",  "Covered",                         ""),
    ("Phase 10: Scheduling",         "✅ Post notifications",                    "Minor: Auto-publisher",            ""),
    ("Phase 11: Telegram Tracking",  "❌ Not built",                             "BIGGEST GAP — 4 workflows needed","HIGH"),
    ("Phase 12: Chrome Dashboard",   "N/A (manual)",                            "N/A",                             ""),
]

for i, (phase, coverage, gap, priority) in enumerate(COVERAGE):
    bg = C_ROW_ALT if i % 2 == 0 else "FFFFFF"
    row_data = [phase, coverage, gap, "", "", "", ""]
    for col, val in enumerate(row_data, 1):
        c = ws.cell(row, col, val)
        c.border = thin_border()
        c.font = Font(size=9)
        c.alignment = left()
        if col == 3:
            if "GAP" in gap.upper() or "BIGGEST" in gap.upper():
                c.fill = fill(C_RED)
            elif "Covered" in gap:
                c.fill = fill(C_GREEN)
            elif "Minor" in gap:
                c.fill = fill(C_YELLOW)
            elif "N/A" in gap:
                c.fill = fill(C_ROW_ALT)
            else:
                c.fill = fill(C_YELLOW)
        elif col == 2:
            c.fill = status_fill(coverage)
        else:
            c.fill = fill(bg)
    ws.row_dimensions[row].height = 15
    row += 1

# ── Column widths ─────────────────────────────────────────────────────────────
col_widths = [16, 30, 30, 18, 45, 16, 30]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Freeze top rows (title + blank + section header + col header = row 5)
ws.freeze_panes = "A5"

wb.save(PATH)
print("Done — sheet added to 00_AGENT_COMMAND.xlsx")
